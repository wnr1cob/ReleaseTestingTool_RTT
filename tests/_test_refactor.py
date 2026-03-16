"""
Refactoring validation tests – run directly to confirm all logic fixes.
"""
import os
import sys

# Ensure project root is on the path when run directly
_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _root not in sys.path:
    sys.path.insert(0, _root)

import tempfile
import shutil
import datetime
import traceback

PASS = "\u2705"
FAIL = "\u274c"
WARN = "\u26a0"

errors: list[str] = []


def check(label: str, fn, expected=True):
    try:
        got = fn()
        ok = got == expected
        print(f"{PASS if ok else FAIL}  {label}")
        if not ok:
            print(f"         expected={expected!r}  got={got!r}")
            errors.append(label)
    except Exception as e:
        print(f"{FAIL}  {label}  -> {type(e).__name__}: {e}")
        errors.append(label)
        traceback.print_exc()


# ── utils ────────────────────────────────────────────────────────
print("=== utils ===")
from src.core.systemtestliste.utils import cell_to_str, parse_sw_variant

check("parse_sw_variant MySW_v3", lambda: parse_sw_variant("MySW_v3"), ("MySW", "v3"))
check("parse_sw_variant no variant", lambda: parse_sw_variant("X"), ("X", ""))
check("parse_sw_variant empty", lambda: parse_sw_variant(""), ("", ""))
check("cell_to_str None -> empty", lambda: cell_to_str(None), "")
check("cell_to_str 42.0 -> '42'", lambda: cell_to_str(42.0), "42")
check("cell_to_str 3.14 -> '3.14'", lambda: cell_to_str(3.14), "3.14")
check("cell_to_str whitespace stripped", lambda: cell_to_str("  hi  "), "hi")
check("cell_to_str date", lambda: cell_to_str(datetime.date(2024, 1, 1)), "2024-01-01")
check("cell_to_str datetime", lambda: cell_to_str(datetime.datetime(2024, 1, 1, 10, 0, 0)), "2024-01-01T10:00:00")
check("cell_to_str int passthrough", lambda: cell_to_str(7), "7")

# ── excel_reader ─────────────────────────────────────────────────
print()
print("=== excel_reader ===")
from src.core.systemtestliste.excel_reader import filter_hilts_rows, find_header_row

rows = [
    ["Random", "stuff"],
    ["ID", "Description", "Result", "Function"],
    ["1", "HILTS task A", "Passed", "f1"],
    ["2", "Regular row", "Failed", "f2"],
    ["3", "HILTS task B", "Error9", "f3"],  # digit in Result -> stripped to "Error"
    [],  # fully empty -> skipped
]
idx, hdr, cm = find_header_row(rows)
check("header found at row 1", lambda: idx, 1)
check("ID -> col 0", lambda: cm["ID"], 0)
check("Description -> col 1", lambda: cm["Description"], 1)
check("Result -> col 2", lambda: cm["Result"], 2)

data = filter_hilts_rows(rows, idx, cm)
check("2 HILTS rows found", lambda: len(data), 2)
check("first HILTS desc correct", lambda: data[0][1], "HILTS task A")
check("first HILTS result correct", lambda: data[0][2], "Passed")
check("digit stripped from Error9 -> Error", lambda: data[1][2], "Error")
check("non-HILTS row excluded", lambda: any("Regular" in r[1] for r in data), False)
check("empty row skipped", lambda: any(not any(c for c in r) for r in data), False)

try:
    find_header_row([["foo", "bar"], ["baz", "qux"]])
    errors.append("no ValueError for headerless sheet")
    print(f"{FAIL}  ValueError not raised for headerless sheet")
except ValueError:
    print(f"{PASS}  ValueError raised for headerless sheet")

# ── file_copier ──────────────────────────────────────────────────
print()
print("=== file_copier ===")
from src.core.pdf_analyzer.file_copier import copy_pdfs

tmp = tempfile.mkdtemp()
src_dir = os.path.join(tmp, "src")
dst_dir = os.path.join(tmp, "dst")
os.makedirs(src_dir)
files = []
for name in ["A-R.pdf", "B-R.pdf", "A-R.pdf"]:
    p = os.path.join(src_dir, name)
    open(p, "wb").write(b"x")
    files.append(p)

r = copy_pdfs(files, dst_dir, mode="copy_duplicates")
check("copy_duplicates copied=3", lambda: r["copied"], 3)
check("copy_duplicates skipped=0", lambda: r["skipped"], 0)
check("_Dup file exists", lambda: os.path.isfile(os.path.join(dst_dir, "A-R_Dup.pdf")), True)

r2 = copy_pdfs(files[:1], dst_dir, mode="ignore_duplicates")
check("ignore_duplicates skipped=1", lambda: r2["skipped"], 1)
check("ignore_duplicates copied=0", lambda: r2["copied"], 0)

calls: list = []
copy_pdfs(files[:2], dst_dir, "ignore_duplicates", lambda p, t: calls.append((p, t)))
check("progress called twice", lambda: len(calls), 2)
check("progress (1,2) then (2,2)", lambda: calls, [(1, 2), (2, 2)])
shutil.rmtree(tmp)

# ── pdf_matcher ──────────────────────────────────────────────────
print()
print("=== pdf_matcher ===")
from src.core.systemtestliste.pdf_matcher import build_pdf_index, match_all_rows, match_pdf_result

tmp = tempfile.mkdtemp()
for name in ["ABC-Module.pdf", "XYZ-Result.pdf"]:
    open(os.path.join(tmp, name), "wb").write(b"x")
idx2 = build_pdf_index(tmp)
check("2 PDFs indexed", lambda: len(idx2), 2)
check("empty index -> No report", lambda: match_pdf_result("x", []), "No report")
check("low score -> No report", lambda: match_pdf_result("totally different description", idx2), "No report")

empty_results = match_all_rows([], idx2)
check("empty rows -> []", lambda: empty_results, [])

prog_calls: list = []
match_all_rows([["1", "desc", "P"]], idx2, on_progress=lambda p, t, r: prog_calls.append((p, t, r)))
check("progress callback (1,1,...)", lambda: prog_calls[0][:2], (1, 1))
check("progress result is string", lambda: isinstance(prog_calls[0][2], str), True)
shutil.rmtree(tmp)

# ── report_writer ────────────────────────────────────────────────
print()
print("=== report_writer ===")
from src.core.systemtestliste.report_writer import write_stl_helper

import openpyxl

tmp = tempfile.mkdtemp()
path = write_stl_helper(
    tmp,
    [["A", "B", "C"], ["D", "E", "F"]],
    ["Passed", "Failed"],
    ["C1", "C2", "C3"],
    "TS1",
)
wb = openpyxl.load_workbook(path)
ws = wb.active
check("header + 2 data rows = 3 total", lambda: ws.max_row, 3)
check("PDFResult header in col 4", lambda: ws.cell(1, 4).value, "PDFResult")
check("row2 col4 = Passed", lambda: ws.cell(2, 4).value, "Passed")
check("row3 col4 = Failed", lambda: ws.cell(3, 4).value, "Failed")

# Length mismatch must raise ValueError (bug fix)
try:
    write_stl_helper(tmp, [["A"], ["B"], ["C"]], ["X"], ["C1"], "TS2")
    errors.append("no ValueError on length mismatch")
    print(f"{FAIL}  ValueError not raised on length mismatch")
except ValueError as e:
    print(f"{PASS}  ValueError raised for mismatched lengths: {e}")

shutil.rmtree(tmp)

# ── sheet_helpers ─────────────────────────────────────────────────
print()
print("=== sheet_helpers ===")
from src.core.pdf_analyzer.sheet_helpers import _extract_module

check("ABC-Report.pdf -> ABC", lambda: _extract_module("ABC-Report.pdf"), "ABC")
check("NoPrefix.pdf -> Unknown", lambda: _extract_module("NoPrefix.pdf"), "Unknown")
check("empty string -> Unknown", lambda: _extract_module(""), "Unknown")
check("-StartDash.pdf -> Unknown (was '' before fix)", lambda: _extract_module("-StartDash.pdf"), "Unknown")
check("A-B-C.pdf -> A (first prefix only)", lambda: _extract_module("A-B-C.pdf"), "A")
check("  -SpacedDash.pdf -> Unknown", lambda: _extract_module("  -SpacedDash.pdf"), "Unknown")

# ── module_separator ─────────────────────────────────────────────
print()
print("=== module_separator ===")
from src.core.pdf_analyzer.module_separator import separate_by_module

tmp = tempfile.mkdtemp()
for name in ["ABC-R1.pdf", "ABC-R2.pdf", "XYZ-R1.pdf", "NoDash.pdf", "-Bad.pdf"]:
    open(os.path.join(tmp, name), "wb").write(b"%PDF")

r3 = separate_by_module(tmp)
check("moved=3", lambda: r3["moved"], 3)
check("skipped=2 (NoDash + -Bad)", lambda: r3["skipped"], 2)
check("modules=['ABC','XYZ']", lambda: r3["modules"], ["ABC", "XYZ"])
check("dest_root ends Module_Separated", lambda: r3["dest_root"].endswith("Module_Separated"), True)

# Duplicate within module folder gets renamed
open(os.path.join(tmp, "ABC-R1.pdf"), "wb").write(b"%PDF-dup")
r4 = separate_by_module(tmp)
abc_files = os.listdir(os.path.join(r4["dest_root"], "ABC"))
check("duplicate renamed with _1 suffix", lambda: any("_1" in f for f in abc_files), True)
shutil.rmtree(tmp)

# ── generate_report ──────────────────────────────────────────────
print()
print("=== generate_report ===")
from src.core.pdf_analyzer.report_generator import generate_report

tmp = tempfile.mkdtemp()
file_results = [
    {"name": "ABC-R.pdf", "result": "Passed"},
    {"name": "XYZ-R.pdf", "result": "Failed"},
    {"name": "-Bad.pdf", "result": "Error"},   # empty module prefix
    {"name": "NoModule.pdf", "result": "Unknown"},
]
path2 = generate_report(tmp, file_results, {"Passed": 1, "Failed": 1, "Error": 1})
wb2 = openpyxl.load_workbook(path2)
check("Execution Summary sheet", lambda: "Execution Summary" in wb2.sheetnames, True)
check("Modules sheet", lambda: "Modules" in wb2.sheetnames, True)

ws2 = wb2["Modules"]
mod_names = [ws2.cell(r, 2).value for r in range(6, ws2.max_row + 1)]
check("'Unknown' module in Modules sheet (was '' before fix)", lambda: "Unknown" in mod_names, True)
check("'ABC' module in Modules sheet", lambda: "ABC" in mod_names, True)

# Zero total edge case (no div-by-zero)
path3 = generate_report(tmp, [], {})
check("empty report created without exception", lambda: os.path.isfile(path3), True)
shutil.rmtree(tmp)

# ── Summary ──────────────────────────────────────────────────────
print()
print("=" * 55)
if errors:
    print(f"{FAIL}  {len(errors)} FAILURE(S):")
    for e in errors:
        print(f"    - {e}")
    sys.exit(1)
else:
    print(f"{PASS}  ALL TESTS PASSED")
