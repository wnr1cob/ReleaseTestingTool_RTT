"""
Excel reading utilities for the SystemTestListe analyser.

Public API
----------
load_sheet_names(filepath)
    Return all sheet names from an .xlsx or .xls file.

read_sheet_data(filepath, sheet_name)
    Read all rows from a specific sheet as ``list[list[str]]``.

find_header_row(all_rows)
    Detect the header row and return ``(row_idx, header_names, col_map)``.

filter_hilts_rows(all_rows, header_row_idx, col_map)
    Return rows where the Description cell contains "HILTS".
"""
import os
import re as _re

from .utils import cell_to_str


# ── Column-detection constants ──────────────────────────────────
HEADER_CANDIDATES = [
    "ID", "Description", "Result", "Comment", "Problem Job",
    "Expert judgement", "Function", "Test Instance", "Safety",
    "ML", "Inspector", "Part of Concept",
]

OUTPUT_COLS = ["ID", "Description", "Result"]


# ═══════════════════════════════════════════════════════════════
# PUBLIC FUNCTIONS
# ═══════════════════════════════════════════════════════════════

def load_sheet_names(filepath: str) -> list[str]:
    """Return all sheet names from an Excel file (``.xlsx`` or ``.xls``)."""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(filepath)
        return wb.sheet_names()
    else:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names


def read_sheet_data(filepath: str, sheet_name: str) -> list[list[str]]:
    """Read all rows from *sheet_name* as a list of string lists.

    Parameters
    ----------
    filepath : str
        Path to the Excel file.
    sheet_name : str
        Name of the sheet to read.

    Returns
    -------
    list[list[str]]
        Every row as a list of cleaned string values.
    """
    ext = os.path.splitext(filepath)[1].lower()
    if ext == ".xls":
        import xlrd
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_name(sheet_name)
        return [
            [cell_to_str(ws.cell_value(r, c)) for c in range(ws.ncols)]
            for r in range(ws.nrows)
        ]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[sheet_name]
        rows = [
            [cell_to_str(v) for v in row]
            for row in ws.iter_rows(values_only=True)
        ]
        wb.close()
        return rows


def find_header_row(
    all_rows: list[list[str]],
    candidates: list[str] | None = None,
    output_cols: list[str] | None = None,
) -> tuple[int, list[str], dict[str, int | None]]:
    """Find the header row by counting matching candidate column names.

    Parameters
    ----------
    all_rows : list[list[str]]
        All rows from the sheet (output of :func:`read_sheet_data`).
    candidates : list[str], optional
        Column name candidates used to identify the header row.
        Defaults to :data:`HEADER_CANDIDATES`.
    output_cols : list[str], optional
        Columns to build a ``col_map`` for.
        Defaults to :data:`OUTPUT_COLS`.

    Returns
    -------
    (header_row_idx, header_names, col_map)
        ``col_map`` maps each output-column name to its 0-based index
        in *all_rows*, or ``None`` when the column is not found.

    Raises
    ------
    ValueError
        When no header row can be identified.
    """
    if candidates is None:
        candidates = HEADER_CANDIDATES
    if output_cols is None:
        output_cols = OUTPUT_COLS

    candidates_lower = [c.lower() for c in candidates]
    best_idx, best_score = 0, 0

    for i, row in enumerate(all_rows):
        cells_lower = [str(c).strip().lower() for c in row]
        score = sum(1 for cand in candidates_lower if cand in cells_lower)
        if score > best_score:
            best_score, best_idx = score, i

    if best_score == 0:
        raise ValueError("Could not find header row in the selected sheet.")

    header_row = [str(c).strip() for c in all_rows[best_idx]]
    header_lower = [c.lower() for c in header_row]

    col_map: dict[str, int | None] = {}
    for col_name in output_cols:
        try:
            col_map[col_name] = header_lower.index(col_name.lower())
        except ValueError:
            col_map[col_name] = None

    return best_idx, header_row, col_map


def filter_hilts_rows(
    all_rows: list[list[str]],
    header_row_idx: int,
    col_map: dict[str, int | None],
    output_cols: list[str] | None = None,
) -> list[list[str]]:
    """Return rows (after the header) where Description contains "HILTS".

    Parameters
    ----------
    all_rows : list[list[str]]
        All rows from the sheet.
    header_row_idx : int
        0-based index of the header row.
    col_map : dict
        Maps column names to source indices (from :func:`find_header_row`).
    output_cols : list[str], optional
        Columns to extract.  Defaults to :data:`OUTPUT_COLS`.

    Returns
    -------
    list[list[str]]
        Each entry contains only the *output_cols* values.
        The ``Result`` column has its digit characters stripped.
    """
    if output_cols is None:
        output_cols = OUTPUT_COLS

    desc_idx = col_map.get("Description")
    data_rows: list[list[str]] = []

    for row in all_rows[header_row_idx + 1:]:
        if not any(str(c).strip() for c in row):
            continue  # skip fully empty rows
        desc_val = (
            str(row[desc_idx]).strip()
            if desc_idx is not None and desc_idx < len(row)
            else ""
        )
        if "HILTS" in desc_val.upper():
            extracted = []
            for col_name in output_cols:
                idx = col_map.get(col_name)
                val = (
                    str(row[idx]).strip()
                    if idx is not None and idx < len(row)
                    else ""
                )
                # Strip digits from the Result column
                if col_name == "Result" and val:
                    val = _re.sub(r"\d", "", val).strip()
                extracted.append(val)
            data_rows.append(extracted)

    return data_rows
