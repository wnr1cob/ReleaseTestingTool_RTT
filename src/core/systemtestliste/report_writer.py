"""\nOutput Excel creation for the SystemTestListe analyser.\n\nPublic API\n----------\nwrite_stl_helper(output_dir, data_rows, pdf_results, output_cols, timestamp, ...)\n    Write the ``STL_Helper_<timestamp>.xlsx`` file and return its path.\n\nSheets produced\n---------------\nMainSheet\n    All HILTS rows with all info columns: PDFResult, Report Availability,\n    ResultMatch, Page3_SW, SWMatch, Page3_Variant, VariantMatch, Library Version.\n    Mismatch rows are row-highlighted; individual mismatch cells are\n    additionally highlighted in a brighter shade.\n\nStep 1 — No Reports  (only when some test cases have no matched PDF)\nStep 2 — Result Mismatches  (only when there are result mismatches, no-report rows excluded)\n         SW Mismatches      (only when there are SW mismatches, no-report rows excluded)\n         Variant Mismatches (only when there are variant mismatches, no-report rows excluded)\n         Library — Not Linked (only when library versions are missing, no-report rows excluded)\nStep 3 — Empty tabs are never created / are deleted before saving.\n"""
import os
from datetime import datetime


# -----------------------------------------------------------------
# Colour palette
# -----------------------------------------------------------------
_C_HEADER_FILL        = "1F3864"   # dark navy      -- header background
_C_HEADER_FONT        = "FFFFFF"   # white          -- header text
_C_ALT_ROW            = "EEF2F7"   # light blue-grey -- alternating row
_C_MISMATCH_ROW       = "FFE0E0"   # pale red        -- entire mismatch row tint
_C_MISMATCH_CELL      = "FF6B6B"   # vivid red       -- exact mismatch cell
_C_MATCH_CELL         = "C6EFCE"   # pale green      -- match cell
_C_MISMATCH_TAB_HDR   = "C00000"   # dark red        -- mismatch tab header bg
_C_BORDER             = "B8C4D4"   # light blue-grey border
_C_NOT_LINKED_ROW     = "FFF2CC"   # pale amber  -- "Not linked" row tint
_C_NOT_LINKED_CELL    = "FFC000"   # vivid amber -- "Not linked" cell
_C_NOT_LINKED_TAB_HDR = "7F6000"   # dark amber  -- not-linked tab header
_C_NO_REPORT_ROW      = "F2F2F2"   # light grey  -- no-report row tint
_C_NO_REPORT_CELL     = "A6A6A6"   # mid grey    -- no-report cell


def write_stl_helper(
    output_dir,
    data_rows,
    pdf_results,
    output_cols,
    timestamp=None,
    *,
    sw_name="",
    variant="",
    match_flags=None,
    page3_sw_list=None,
    sw_match_flags=None,
    page3_variant_list=None,
    variant_match_flags=None,
    library_version_list=None,
):
    """Write the STL_Helper Excel file and return its absolute path."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if timestamp is None:
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

    output_path = os.path.join(output_dir, f"STL_Helper_{timestamp}.xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MainSheet"

    # ---- style helpers --------------------------------------------------
    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _border():
        thin = Side(style="thin", color=_C_BORDER)
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _hdr_font():
        return Font(name="Calibri", bold=True, color=_C_HEADER_FONT, size=10)

    def _body_font(bold=False):
        return Font(name="Calibri", bold=bold, size=10)

    def _align(horizontal="left", wrap=False):
        return Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)

    def _align_top(horizontal="left"):
        return Alignment(horizontal=horizontal, vertical="top", wrap_text=False)

    def _apply_header_row(ws_t, headers, bg=_C_HEADER_FILL):
        ws_t.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws_t.cell(row=1, column=col_idx)
            cell.fill      = _fill(bg)
            cell.font      = _hdr_font()
            cell.alignment = _align("center", wrap=False)
            cell.border    = _border()
        ws_t.row_dimensions[1].height = 22

    def _autofit_columns(ws_t):
        for col_cells in ws_t.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    cell_len = len(str(cell.value)) if cell.value is not None else 0
                    max_len = max(max_len, cell_len)
                except Exception:
                    pass
            ws_t.column_dimensions[col_letter].width = min(max(max_len + 6, 14), 150)

    def _autofit_rows(ws_t):
        for row_idx in range(2, ws_t.max_row + 1):
            ws_t.row_dimensions[row_idx].height = 18

    # ---- build MainSheet header -----------------------------------------
    main_headers = list(output_cols) + ["PDFResult", "Report"]
    if match_flags is not None:
        main_headers.append("ResultMatch")
    if page3_sw_list is not None:
        main_headers.append("Page3_SW")
    if sw_match_flags is not None:
        main_headers.append("SWMatch")
    if page3_variant_list is not None:
        main_headers.append("Page3_Variant")
    if variant_match_flags is not None:
        main_headers.append("VariantMatch")
    if library_version_list is not None:
        main_headers.append("Library Version")

    _apply_header_row(ws, main_headers)

    col_idx_map = {name: i + 1 for i, name in enumerate(main_headers)}

    if len(data_rows) != len(pdf_results):
        raise ValueError(
            f"data_rows ({len(data_rows)}) and pdf_results ({len(pdf_results)}) "
            "must have the same length."
        )

    # ---- pre-create reusable style objects --------------------------------
    # Creating a new PatternFill / Font / Border / Alignment per cell is expensive
    # (each is a small Python object). Pre-creating one of each kind, then
    # assigning the *same* object to every qualifying cell, cuts allocations
    # from O(rows × cols) down to O(distinct_styles).
    _pre_border       = _border()
    _pre_body_font    = _body_font()
    _pre_bold_font    = _body_font(bold=True)
    _pre_hdr_font     = _hdr_font()
    _pre_align_left   = _align("left")
    _pre_align_center = _align("center")
    _pre_align_top    = _align_top()

    _pre_fill_mm_row   = _fill(_C_MISMATCH_ROW)
    _pre_fill_nl_row   = _fill(_C_NOT_LINKED_ROW)
    _pre_fill_nr_row   = _fill(_C_NO_REPORT_ROW)
    _pre_fill_alt_row  = _fill(_C_ALT_ROW)
    _pre_fill_mm_cell  = _fill(_C_MISMATCH_CELL)
    _pre_fill_ok_cell  = _fill(_C_MATCH_CELL)
    _pre_fill_nl_cell  = _fill(_C_NOT_LINKED_CELL)
    _pre_fill_nr_cell  = _fill(_C_NO_REPORT_CELL)

    # ---- write MainSheet data rows -------------------------------------
    for i, (row, pdf_result) in enumerate(zip(data_rows, pdf_results)):
        excel_row_idx = i + 2

        no_report        = pdf_results[i] == "No report"
        result_mismatch  = (match_flags is not None) and (not match_flags[i])
        sw_mismatch      = (sw_match_flags is not None) and (not sw_match_flags[i])
        variant_mismatch = (variant_match_flags is not None) and (not variant_match_flags[i])
        lib_not_linked   = (library_version_list is not None) and (not library_version_list[i])
        any_mismatch     = result_mismatch or sw_mismatch or variant_mismatch

        data = list(row) + [pdf_result, "✗ No Report" if no_report else "✓ Available"]
        if match_flags is not None:
            data.append("PASS" if match_flags[i] else "FAIL")
        if page3_sw_list is not None:
            data.append(page3_sw_list[i])
        if sw_match_flags is not None:
            data.append("MATCH" if sw_match_flags[i] else "MISMATCH")
        if page3_variant_list is not None:
            data.append(page3_variant_list[i])
        if variant_match_flags is not None:
            data.append("MATCH" if variant_match_flags[i] else "MISMATCH")
        if library_version_list is not None:
            lib_val = library_version_list[i] if library_version_list[i] else "Not linked"
            data.append(lib_val)

        ws.append(data)

        row_fill = _pre_fill_mm_row  if any_mismatch \
                   else (_pre_fill_nl_row  if lib_not_linked \
                   else (_pre_fill_nr_row  if no_report \
                   else (_pre_fill_alt_row if i % 2 == 1 else None)))

        for col_idx in range(1, len(data) + 1):
            cell           = ws.cell(row=excel_row_idx, column=col_idx)
            cell.font      = _pre_body_font
            cell.alignment = _pre_align_top if col_idx <= 2 else _pre_align_left
            cell.border    = _pre_border
            if row_fill:
                cell.fill = row_fill

        # Paint match/mismatch indicator cells
        def _paint(col_name, is_mismatch):
            if col_name not in col_idx_map:
                return
            c = ws.cell(row=excel_row_idx, column=col_idx_map[col_name])
            c.fill = _pre_fill_mm_cell if is_mismatch else _pre_fill_ok_cell
            c.font = _pre_bold_font    if is_mismatch else _pre_body_font

        if match_flags is not None:
            _paint("ResultMatch", result_mismatch)
            if result_mismatch:
                ws.cell(row=excel_row_idx,
                        column=col_idx_map["PDFResult"]).fill = _pre_fill_mm_cell
                stl_res_col = col_idx_map.get("Result")
                if stl_res_col:
                    ws.cell(row=excel_row_idx,
                            column=stl_res_col).fill = _pre_fill_mm_cell

        if sw_match_flags is not None:
            _paint("SWMatch", sw_mismatch)
            if sw_mismatch and "Page3_SW" in col_idx_map:
                ws.cell(row=excel_row_idx,
                        column=col_idx_map["Page3_SW"]).fill = _pre_fill_mm_cell

        if variant_match_flags is not None:
            _paint("VariantMatch", variant_mismatch)
            if variant_mismatch and "Page3_Variant" in col_idx_map:
                ws.cell(row=excel_row_idx,
                        column=col_idx_map["Page3_Variant"]).fill = _pre_fill_mm_cell

        if library_version_list is not None and lib_not_linked:
            lib_col = col_idx_map.get("Library Version")
            if lib_col:
                c = ws.cell(row=excel_row_idx, column=lib_col)
                c.fill = _pre_fill_nl_cell
                c.font = _pre_bold_font

        # Paint Report Availability cell
        rep_col = col_idx_map.get("Report")
        if rep_col:
            c = ws.cell(row=excel_row_idx, column=rep_col)
            if no_report:
                c.fill = _pre_fill_nr_cell
                c.font = _pre_bold_font
            else:
                c.fill = _pre_fill_ok_cell

    ws.freeze_panes = "A2"
    _autofit_columns(ws)
    _autofit_rows(ws)

    # ---- mismatch detail tab helper ------------------------------------
    def _write_mismatch_tab(tab_title, parameter, mm_indices, expected_vals, obtained_vals):
        if not mm_indices:
            return
        ws_tab = wb.create_sheet(title=tab_title)
        headers = ["Test Case ID", "Test Case Name", "Parameter",
                   "Expected Value", "Obtained Value (PDF)"]
        _apply_header_row(ws_tab, headers, bg=_C_MISMATCH_TAB_HDR)

        for row_num, (row_i, exp, obt) in enumerate(
                zip(mm_indices, expected_vals, obtained_vals), start=2):
            base = data_rows[row_i]
            tc_id   = base[0] if len(base) > 0 else ""
            tc_name = base[1] if len(base) > 1 else ""
            ws_tab.append([tc_id, tc_name, parameter, exp, obt])
            for col_idx in range(1, len(headers) + 1):
                cell           = ws_tab.cell(row=row_num, column=col_idx)
                cell.font      = _pre_bold_font if col_idx in (4, 5) else _pre_body_font
                cell.alignment = _pre_align_top if col_idx <= 2 else _pre_align_left
                cell.border    = _pre_border
            ws_tab.cell(row=row_num, column=4).fill = _pre_fill_ok_cell
            ws_tab.cell(row=row_num, column=5).fill = _pre_fill_mm_cell
            if row_num % 2 == 0:
                for col_idx in range(1, 4):
                    cell = ws_tab.cell(row=row_num, column=col_idx)
                    rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else ""
                    if rgb in ("00000000", "FFFFFFFF", "00FFFFFF", ""):
                        cell.fill = _pre_fill_alt_row

        ws_tab.freeze_panes = "A2"
        _autofit_columns(ws_tab)
        _autofit_rows(ws_tab)

        summary_row = ws_tab.max_row + 2
        c = ws_tab.cell(row=summary_row, column=1,
                        value=f"Total mismatches: {len(mm_indices)}")
        c.font = _body_font(bold=True)

    # ---- No Reports tab (items where no PDF was matched) ---------------
    nr_idx = [i for i, r in enumerate(pdf_results) if r == "No report"]
    if nr_idx:
        ws_nr = wb.create_sheet(title="No Reports")
        nr_headers = ["Test Case ID", "Test Case Name"]
        _apply_header_row(ws_nr, nr_headers, bg=_C_MISMATCH_TAB_HDR)
        for row_num, row_i in enumerate(nr_idx, start=2):
            base    = data_rows[row_i]
            tc_id   = base[0] if len(base) > 0 else ""
            tc_name = base[1] if len(base) > 1 else ""
            ws_nr.append([tc_id, tc_name])
            for col_idx in range(1, len(nr_headers) + 1):
                cell           = ws_nr.cell(row=row_num, column=col_idx)
                cell.font      = _pre_body_font
                cell.alignment = _pre_align_top if col_idx <= 2 else _pre_align_left
                cell.border    = _pre_border
            if row_num % 2 == 0:
                for col_idx in range(1, len(nr_headers) + 1):
                    ws_nr.cell(row=row_num, column=col_idx).fill = _pre_fill_alt_row
        ws_nr.freeze_panes = "A2"
        _autofit_columns(ws_nr)
        _autofit_rows(ws_nr)
        summary_row = ws_nr.max_row + 2
        c = ws_nr.cell(row=summary_row, column=1,
                       value=f"Total no reports: {len(nr_idx)}")
        c.font = _body_font(bold=True)

    # ---- create mismatch tabs ------------------------------------------
    if match_flags is not None:
        stl_res_idx = output_cols.index("Result") if "Result" in output_cols else 2
        mm_idx = [i for i, ok in enumerate(match_flags)
                  if not ok and pdf_results[i] != "No report"]
        _write_mismatch_tab(
            "Result Mismatches", "Result", mm_idx,
            [data_rows[i][stl_res_idx] for i in mm_idx],
            [pdf_results[i] for i in mm_idx],
        )

    if sw_match_flags is not None and page3_sw_list is not None:
        mm_idx = [i for i, ok in enumerate(sw_match_flags)
                  if not ok and pdf_results[i] != "No report"]
        _write_mismatch_tab(
            "SW Mismatches", "SW Version", mm_idx,
            [sw_name] * len(mm_idx),
            [page3_sw_list[i] for i in mm_idx],
        )

    if variant_match_flags is not None and page3_variant_list is not None:
        mm_idx = [i for i, ok in enumerate(variant_match_flags)
                  if not ok and pdf_results[i] != "No report"]
        _write_mismatch_tab(
            "Variant Mismatches", "Variant", mm_idx,
            [variant] * len(mm_idx),
            [page3_variant_list[i] for i in mm_idx],
        )

    # ---- Library — Not Linked tab --------------------------------------
    if library_version_list is not None:
        nl_idx = [i for i, v in enumerate(library_version_list)
                  if not v and pdf_results[i] != "No report"]
        if nl_idx:
            ws_nl = wb.create_sheet(title="Library — Not Linked")
            nl_headers = ["Test Case ID", "Test Case Name", "Library Version"]
            _apply_header_row(ws_nl, nl_headers, bg=_C_NOT_LINKED_TAB_HDR)
            for row_num, row_i in enumerate(nl_idx, start=2):
                base = data_rows[row_i]
                tc_id   = base[0] if len(base) > 0 else ""
                tc_name = base[1] if len(base) > 1 else ""
                ws_nl.append([tc_id, tc_name, "Not linked"])
                for col_idx in range(1, len(nl_headers) + 1):
                    cell           = ws_nl.cell(row=row_num, column=col_idx)
                    cell.font      = _pre_bold_font if col_idx == 3 else _pre_body_font
                    cell.alignment = _pre_align_top if col_idx <= 2 else _pre_align_left
                    cell.border    = _pre_border
                ws_nl.cell(row=row_num, column=3).fill = _pre_fill_nl_cell
                if row_num % 2 == 0:
                    for col_idx in range(1, 3):
                        cell = ws_nl.cell(row=row_num, column=col_idx)
                        rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else ""
                        if rgb in ("00000000", "FFFFFFFF", "00FFFFFF", ""):
                            cell.fill = _pre_fill_alt_row
            ws_nl.freeze_panes = "A2"
            _autofit_columns(ws_nl)
            _autofit_rows(ws_nl)
            summary_row = ws_nl.max_row + 2
            c = ws_nl.cell(row=summary_row, column=1,
                           value=f"Total not linked: {len(nl_idx)}")
            c.font = _body_font(bold=True)

    wb.save(output_path)
    return output_path
