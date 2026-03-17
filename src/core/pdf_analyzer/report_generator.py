"""
Report Generator – create an Excel summary of PDF analysis results.

Produces ``RTT_Execution_YYYYMMDDHHMMSS.xlsx`` with:
* Header banner with title and timestamp
* Stat cards in a single row: Passed | Failed | Error | Undefined | Total
* Detail table with auto-width filename column and status-specific styling
"""

import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font

from src.core.pdf_analyzer.styles import (
    _NO_BORDER, _THIN_BORDER,
    _TITLE_FILL, _DARK_FILL, _SECTION_FILL, _TABLE_HEADER_FILL,
    _ROW_EVEN, _ROW_ODD, _TOTAL_FILL,
    _STATUS_FILLS, _CARD_FILLS,
    _TITLE_FONT, _SUBTITLE_FONT,
    _CARD_LABEL_FONT, _CARD_VALUE_FONT, _CARD_PCT_FONT,
    _TABLE_HEADER_FONT, _BODY_FONT, _BODY_BOLD, _STATUS_FONTS,
    _CENTER, _LEFT_CENTER, _RIGHT_CENTER,
)
from src.core.pdf_analyzer.sheet_helpers import (
    _fill_range,
    _auto_col_width,
    _extract_module,
)


# ═══════════════════════════════════════════════════════════════
# PUBLIC API
# ═══════════════════════════════════════════════════════════════

def generate_report(
    output_dir: str,
    file_results: list[dict[str, str]],
    results_count: dict[str, int],
    file_summaries: list[dict[str, str]] | None = None,
) -> str:
    """
    Create the RTT Execution Excel report.

    Parameters
    ----------
    output_dir : str
        Directory where the file will be saved.
    file_results : list[dict]
        ``[{"name": "file.pdf", "result": "Passed"}, ...]``
    results_count : dict[str, int]
        ``{"Passed": 5, "Failed": 2, "Error": 1, "Undefined": 0}``
    file_summaries : list[dict], optional
        ``[{"original": "...", "final": "...", "action": "Renamed"|"As-Is", "result": "..."}]``
        When provided, a **File Summary** sheet is added to the workbook.

    Returns
    -------
    str   Absolute path to the saved ``.xlsx`` file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Execution Summary"
    ws.sheet_properties.tabColor = "00D4FF"
    ws.sheet_view.showGridLines = False

    total = sum(results_count.values())
    categories = ["Passed", "Failed", "Error", "Undefined", "Total"]
    values = [results_count.get(c, 0) for c in categories[:-1]] + [total]
    pcts = [(v / total * 100) if total else 0 for v in values]
    pcts[-1] = 100.0

    # Columns: A=margin  B=#  C=Filename  D=Status  E-F=extra cards  G=margin
    # For the stat cards we need 5 columns (B-F), so add E and F
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 3

    # ───────────────────────────────────────────────────────────
    # ROW 1-3 : TITLE BANNER
    # ───────────────────────────────────────────────────────────
    _fill_range(ws, 1, 3, 1, 7, _TITLE_FILL)
    ws.row_dimensions[1].height = 8

    ws.merge_cells("B2:F2")
    c = ws["B2"]
    c.value = "📊  RTT Execution Summary"
    c.font = _TITLE_FONT
    c.alignment = _LEFT_CENTER
    ws.row_dimensions[2].height = 40

    ws.merge_cells("B3:F3")
    c = ws["B3"]
    c.value = f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
    c.font = _SUBTITLE_FONT
    c.alignment = _LEFT_CENTER
    ws.row_dimensions[3].height = 22

    # ───────────────────────────────────────────────────────────
    # ROW 4-8 : STAT CARDS  (all 5 in one row: B-F)
    # ───────────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 6
    _fill_range(ws, 4, 4, 1, 7, _DARK_FILL)

    ws.row_dimensions[5].height = 18   # label
    ws.row_dimensions[6].height = 34   # value
    ws.row_dimensions[7].height = 18   # percentage
    ws.row_dimensions[8].height = 8    # spacer
    _fill_range(ws, 5, 8, 1, 7, _DARK_FILL)

    card_cols = [2, 3, 4, 5, 6]  # B, C, D, E, F
    for i, col in enumerate(card_cols):
        cat = categories[i]
        fill = _CARD_FILLS[cat]

        lbl = ws.cell(row=5, column=col, value=f"  {cat}")
        lbl.font = _CARD_LABEL_FONT
        lbl.fill = fill
        lbl.alignment = _LEFT_CENTER

        val = ws.cell(row=6, column=col, value=values[i])
        val.font = _CARD_VALUE_FONT
        val.fill = fill
        val.alignment = _CENTER

        pct = ws.cell(row=7, column=col, value=f"  {pcts[i]:.1f}%")
        pct.font = _CARD_PCT_FONT
        pct.fill = fill
        pct.alignment = _LEFT_CENTER

    # ───────────────────────────────────────────────────────────
    # ROW 9 : SECTION HEADER
    # ───────────────────────────────────────────────────────────
    row = 9
    _fill_range(ws, row, row, 1, 7, _SECTION_FILL)
    ws.merge_cells(f"B{row}:F{row}")
    c = ws.cell(row=row, column=2, value="  📄  PDF File Details")
    c.font = Font(name="Segoe UI", size=13, bold=True, color="FFFFFF")
    c.fill = _SECTION_FILL
    c.alignment = _LEFT_CENTER
    ws.row_dimensions[row].height = 32
    row += 1

    # ───────────────────────────────────────────────────────────
    # ROW 10 : TABLE HEADER  (# | Filename | Status)
    # ───────────────────────────────────────────────────────────
    # Merge C-E for filename column in table area
    ws.merge_cells(f"C{row}:E{row}")
    for col, txt in [(2, "#"), (3, "PDF Filename"), (6, "Status")]:
        c = ws.cell(row=row, column=col, value=txt)
        c.font = _TABLE_HEADER_FONT
        c.fill = _TABLE_HEADER_FILL
        c.border = _THIN_BORDER
        c.alignment = _CENTER
    # Fill merged header cells
    for mc in [4, 5]:
        ws.cell(row=row, column=mc).fill = _TABLE_HEADER_FILL
        ws.cell(row=row, column=mc).border = _THIN_BORDER
    ws.row_dimensions[row].height = 28
    row += 1

    # ───────────────────────────────────────────────────────────
    # DATA ROWS
    # ───────────────────────────────────────────────────────────
    sorted_results = sorted(file_results, key=lambda e: e["name"])
    max_name_len = len("PDF Filename")

    for idx, entry in enumerate(sorted_results, start=1):
        status = entry["result"]
        bg = _ROW_EVEN if idx % 2 == 0 else _ROW_ODD

        # Merge C-E for filename
        ws.merge_cells(f"C{row}:E{row}")

        # #
        c = ws.cell(row=row, column=2, value=idx)
        c.font = _BODY_FONT
        c.border = _THIN_BORDER
        c.alignment = _CENTER
        c.fill = bg

        # Filename
        name = entry["name"]
        max_name_len = max(max_name_len, len(name))
        c = ws.cell(row=row, column=3, value=name)
        c.font = _BODY_FONT
        c.border = _THIN_BORDER
        c.alignment = _LEFT_CENTER
        c.fill = bg
        # Fill merged cells
        for mc in [4, 5]:
            ws.cell(row=row, column=mc).fill = bg
            ws.cell(row=row, column=mc).border = _THIN_BORDER

        # Status
        c = ws.cell(row=row, column=6, value=status)
        c.font = _STATUS_FONTS.get(status, _BODY_BOLD)
        c.border = _THIN_BORDER
        c.alignment = _CENTER
        c.fill = _STATUS_FILLS.get(status, bg)

        ws.row_dimensions[row].height = 22
        row += 1

    # ───────────────────────────────────────────────────────────
    # FOOTER ROW
    # ───────────────────────────────────────────────────────────
    ws.merge_cells(f"B{row}:E{row}")
    c = ws.cell(row=row, column=2, value=f"  Total: {total} file(s)")
    c.font = Font(name="Segoe UI", size=10, bold=True, color="01579B")
    c.fill = _TOTAL_FILL
    c.border = _THIN_BORDER
    c.alignment = _LEFT_CENTER
    for mc in [3, 4, 5]:
        ws.cell(row=row, column=mc).fill = _TOTAL_FILL
        ws.cell(row=row, column=mc).border = _THIN_BORDER

    passed_count = results_count.get("Passed", 0)
    pass_rate = (passed_count / total * 100) if total else 0
    c = ws.cell(row=row, column=6, value=f"Pass: {pass_rate:.1f}%")
    c.font = Font(name="Segoe UI", size=10, bold=True, color="1B5E20")
    c.fill = _TOTAL_FILL
    c.border = _THIN_BORDER
    c.alignment = _CENTER
    ws.row_dimensions[row].height = 26

    # ───────────────────────────────────────────────────────────
    # AUTO-WIDTH for filename columns (C-E merged = set C width)
    # ───────────────────────────────────────────────────────────
    # Combined width of C+D+E should fit the longest name
    name_width = max_name_len + 4  # padding
    # Distribute across C, D, E (keep them equal so merge looks right)
    per_col = max(name_width / 3, 16)
    ws.column_dimensions["C"].width = per_col
    ws.column_dimensions["D"].width = per_col
    ws.column_dimensions["E"].width = per_col

    # ───────────────────────────────────────────────────────────
    # FREEZE & PRINT
    # ───────────────────────────────────────────────────────────
    ws.freeze_panes = "B11"   # Freeze below table header row
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5

    # ───────────────────────────────────────────────────────────
    # MODULES TAB
    # ───────────────────────────────────────────────────────────
    _build_modules_sheet(wb, file_results)

    # ───────────────────────────────────────────────────────────
    # FILE SUMMARY TAB
    # ───────────────────────────────────────────────────────────
    if file_summaries:
        _build_file_summary_sheet(wb, file_summaries)

    # ───────────────────────────────────────────────────────────
    # SAVE
    # ───────────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    filename = f"RTT_Execution_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)

    return filepath


# ═══════════════════════════════════════════════════════════════
# MODULES SHEET
# ═══════════════════════════════════════════════════════════════

def _build_modules_sheet(wb: Workbook, file_results: list[dict[str, str]]):
    """
    Add a **Modules** worksheet that breaks down results per module.

    Each module row shows: Total tests, Passed (#, %), Failed (#, %),
    Error (#, %), Undefined (#, %).
    """
    from collections import defaultdict

    # ── Aggregate data ─────────────────────────────────────────
    module_data: dict[str, dict[str, int]] = defaultdict(
        lambda: {"Total": 0, "Passed": 0, "Failed": 0, "Error": 0, "Undefined": 0}
    )
    for entry in file_results:
        mod = _extract_module(entry["name"])
        status = entry["result"]
        module_data[mod]["Total"] += 1
        if status in module_data[mod]:
            module_data[mod][status] += 1

    modules_sorted = sorted(module_data.keys())

    # ── Create sheet ───────────────────────────────────────────
    ms = wb.create_sheet(title="Modules")
    ms.sheet_properties.tabColor = "FFC107"
    ms.sheet_view.showGridLines = False

    # Column widths:
    # A=margin  B=Module  C=Total  D=Passed  E=P%  F=Failed  G=F%
    # H=Error   I=E%      J=Undefined  K=U%   L=margin
    ms.column_dimensions["A"].width = 3
    ms.column_dimensions["B"].width = 30   # will auto-size later
    ms.column_dimensions["C"].width = 10
    ms.column_dimensions["D"].width = 10
    ms.column_dimensions["E"].width = 10
    ms.column_dimensions["F"].width = 10
    ms.column_dimensions["G"].width = 10
    ms.column_dimensions["H"].width = 10
    ms.column_dimensions["I"].width = 10
    ms.column_dimensions["J"].width = 13
    ms.column_dimensions["K"].width = 10
    ms.column_dimensions["L"].width = 3

    last_col = 12  # L

    # ───────────────────────────────────────────────────────────
    # TITLE BANNER  (rows 1-3)
    # ───────────────────────────────────────────────────────────
    _fill_range(ms, 1, 3, 1, last_col, _TITLE_FILL)
    ms.row_dimensions[1].height = 8

    ms.merge_cells("B2:K2")
    c = ms["B2"]
    c.value = "📦  Module Breakdown"
    c.font = _TITLE_FONT
    c.alignment = _LEFT_CENTER
    ms.row_dimensions[2].height = 40

    ms.merge_cells("B3:K3")
    c = ms["B3"]
    c.value = f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
    c.font = _SUBTITLE_FONT
    c.alignment = _LEFT_CENTER
    ms.row_dimensions[3].height = 22

    # ───────────────────────────────────────────────────────────
    # SPACER
    # ───────────────────────────────────────────────────────────
    ms.row_dimensions[4].height = 8
    _fill_range(ms, 4, 4, 1, last_col, _DARK_FILL)

    # ───────────────────────────────────────────────────────────
    # TABLE HEADER  (row 5)
    # ───────────────────────────────────────────────────────────
    headers = [
        (2, "Module"),
        (3, "Total"),
        (4, "Passed"), (5, "P %"),
        (6, "Failed"), (7, "F %"),
        (8, "Error"),  (9, "E %"),
        (10, "Undefined"), (11, "U %"),
    ]
    for col, txt in headers:
        c = ms.cell(row=5, column=col, value=txt)
        c.font = _TABLE_HEADER_FONT
        c.fill = _TABLE_HEADER_FILL
        c.border = _THIN_BORDER
        c.alignment = _CENTER
    ms.row_dimensions[5].height = 28

    # ───────────────────────────────────────────────────────────
    # DATA ROWS
    # ───────────────────────────────────────────────────────────
    row = 6
    max_mod_len = len("Module")

    for idx, mod in enumerate(modules_sorted, start=1):
        d = module_data[mod]
        mod_total = d["Total"]
        bg = _ROW_EVEN if idx % 2 == 0 else _ROW_ODD

        max_mod_len = max(max_mod_len, len(mod))

        def _pct(count):
            return f"{(count / mod_total * 100):.1f}%" if mod_total else "0.0%"

        row_data = [
            (2, mod,          _BODY_BOLD,                           _LEFT_CENTER, bg),
            (3, mod_total,    _BODY_BOLD,                           _CENTER,      bg),
            (4, d["Passed"],  _STATUS_FONTS["Passed"],              _CENTER,      _STATUS_FILLS["Passed"]),
            (5, _pct(d["Passed"]),  _STATUS_FONTS["Passed"],        _CENTER,      _STATUS_FILLS["Passed"]),
            (6, d["Failed"],  _STATUS_FONTS["Failed"],              _CENTER,      _STATUS_FILLS["Failed"]),
            (7, _pct(d["Failed"]),  _STATUS_FONTS["Failed"],        _CENTER,      _STATUS_FILLS["Failed"]),
            (8, d["Error"],   _STATUS_FONTS["Error"],               _CENTER,      _STATUS_FILLS["Error"]),
            (9, _pct(d["Error"]),   _STATUS_FONTS["Error"],         _CENTER,      _STATUS_FILLS["Error"]),
            (10, d["Undefined"], _STATUS_FONTS["Undefined"],        _CENTER,      _STATUS_FILLS["Undefined"]),
            (11, _pct(d["Undefined"]), _STATUS_FONTS["Undefined"],  _CENTER,      _STATUS_FILLS["Undefined"]),
        ]

        for col, val, font, align, fill in row_data:
            c = ms.cell(row=row, column=col, value=val)
            c.font = font
            c.border = _THIN_BORDER
            c.alignment = align
            c.fill = fill

        ms.row_dimensions[row].height = 24
        row += 1

    # ───────────────────────────────────────────────────────────
    # FOOTER – grand total row
    # ───────────────────────────────────────────────────────────
    grand = {"Total": 0, "Passed": 0, "Failed": 0, "Error": 0, "Undefined": 0}
    for d in module_data.values():
        for k in grand:
            grand[k] += d[k]
    gt = grand["Total"]

    def _gpct(count):
        return f"{(count / gt * 100):.1f}%" if gt else "0.0%"

    footer_data = [
        (2, "TOTAL",          Font(name="Segoe UI", size=10, bold=True, color="01579B"), _LEFT_CENTER),
        (3, gt,               Font(name="Segoe UI", size=10, bold=True, color="01579B"), _CENTER),
        (4, grand["Passed"],  _STATUS_FONTS["Passed"],  _CENTER),
        (5, _gpct(grand["Passed"]),  _STATUS_FONTS["Passed"],  _CENTER),
        (6, grand["Failed"],  _STATUS_FONTS["Failed"],  _CENTER),
        (7, _gpct(grand["Failed"]),  _STATUS_FONTS["Failed"],  _CENTER),
        (8, grand["Error"],   _STATUS_FONTS["Error"],   _CENTER),
        (9, _gpct(grand["Error"]),   _STATUS_FONTS["Error"],   _CENTER),
        (10, grand["Undefined"], _STATUS_FONTS["Undefined"], _CENTER),
        (11, _gpct(grand["Undefined"]), _STATUS_FONTS["Undefined"], _CENTER),
    ]

    for col, val, font, align in footer_data:
        c = ms.cell(row=row, column=col, value=val)
        c.font = font
        c.border = _THIN_BORDER
        c.alignment = align
        c.fill = _TOTAL_FILL

    ms.row_dimensions[row].height = 26

    # ───────────────────────────────────────────────────────────
    # AUTO-WIDTH for Module column
    # ───────────────────────────────────────────────────────────
    ms.column_dimensions["B"].width = max(max_mod_len + 4, 20)

    # ── Freeze & print ────────────────────────────────────────
    ms.freeze_panes = "B6"  # Freeze below header row
    ms.print_options.horizontalCentered = True
    ms.page_margins.left = 0.5
    ms.page_margins.right = 0.5


# ═══════════════════════════════════════════════════════════════
# FILE SUMMARY SHEET
# ═══════════════════════════════════════════════════════════════

def _build_file_summary_sheet(wb: Workbook, file_summaries: list[dict[str, str]]):
    """
    Add a **File Summary** worksheet showing per-file processing details.

    Columns: # | Original Filename | Final Filename | Action | Result
    """
    from openpyxl.styles import PatternFill

    _ACTION_FILL_RENAMED = PatternFill("solid", fgColor="E3F2FD")   # light blue
    _ACTION_FILL_ASIS    = PatternFill("solid", fgColor="F5F5F5")   # light grey
    _ACTION_FONT_RENAMED = Font(name="Segoe UI", size=10, bold=True,  color="0D47A1")
    _ACTION_FONT_ASIS    = Font(name="Segoe UI", size=10, bold=False, color="616161")

    fs = wb.create_sheet(title="File Summary")
    fs.sheet_properties.tabColor = "7B2FF7"
    fs.sheet_view.showGridLines = False

    # Column widths
    fs.column_dimensions["A"].width = 3   # margin
    fs.column_dimensions["B"].width = 6   # #
    fs.column_dimensions["G"].width = 3   # margin
    # C, D, E, F will be auto-sized after data

    last_col = 7  # G

    # ── Title banner (rows 1-3) ─────────────────────────────
    _fill_range(fs, 1, 3, 1, last_col, _TITLE_FILL)
    fs.row_dimensions[1].height = 8

    fs.merge_cells("B2:F2")
    c = fs["B2"]
    c.value = "File Processing Summary"
    c.font = _TITLE_FONT
    c.alignment = _LEFT_CENTER
    fs.row_dimensions[2].height = 40

    fs.merge_cells("B3:F3")
    c = fs["B3"]
    c.value = f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M:%S')}"
    c.font = _SUBTITLE_FONT
    c.alignment = _LEFT_CENTER
    fs.row_dimensions[3].height = 22

    fs.row_dimensions[4].height = 8
    _fill_range(fs, 4, 4, 1, last_col, _DARK_FILL)

    # ── Table header (row 5) ────────────────────────────────
    headers = [
        (2, "#"),
        (3, "Original Filename"),
        (4, "Final Filename"),
        (5, "Action"),
        (6, "Result"),
    ]
    for col, txt in headers:
        c = fs.cell(row=5, column=col, value=txt)
        c.font = _TABLE_HEADER_FONT
        c.fill = _TABLE_HEADER_FILL
        c.border = _THIN_BORDER
        c.alignment = _CENTER
    fs.row_dimensions[5].height = 28

    # ── Data rows ───────────────────────────────────────────
    row = 6
    max_orig_len  = len("Original Filename")
    max_final_len = len("Final Filename")

    sorted_summaries = sorted(file_summaries, key=lambda e: e["final"])

    for idx, entry in enumerate(sorted_summaries, start=1):
        original = entry["original"]
        final    = entry["final"]
        action   = entry["action"]
        result   = entry["result"]
        bg       = _ROW_EVEN if idx % 2 == 0 else _ROW_ODD

        max_orig_len  = max(max_orig_len,  len(original))
        max_final_len = max(max_final_len, len(final))

        action_fill = _ACTION_FILL_RENAMED if action == "Renamed" else _ACTION_FILL_ASIS
        action_font = _ACTION_FONT_RENAMED if action == "Renamed" else _ACTION_FONT_ASIS

        row_data = [
            (2, idx,      _BODY_FONT,                           _CENTER,     bg),
            (3, original, _BODY_FONT,                           _LEFT_CENTER, bg),
            (4, final,    _BODY_BOLD if action == "Renamed" else _BODY_FONT,
                                                                 _LEFT_CENTER, bg),
            (5, action,   action_font,                           _CENTER,     action_fill),
            (6, result,   _STATUS_FONTS.get(result, _BODY_BOLD), _CENTER,     _STATUS_FILLS.get(result, bg)),
        ]

        for col, val, font, align, fill in row_data:
            c = fs.cell(row=row, column=col, value=val)
            c.font = font
            c.border = _THIN_BORDER
            c.alignment = align
            c.fill = fill

        fs.row_dimensions[row].height = 22
        row += 1

    # ── Footer ──────────────────────────────────────────────
    total = len(file_summaries)
    renamed_count = sum(1 for e in file_summaries if e["action"] == "Renamed")

    fs.merge_cells(f"B{row}:D{row}")
    c = fs.cell(row=row, column=2, value=f"  Total: {total} file(s)")
    c.font = Font(name="Segoe UI", size=10, bold=True, color="01579B")
    c.fill = _TOTAL_FILL
    c.border = _THIN_BORDER
    c.alignment = _LEFT_CENTER
    for mc in [3, 4]:
        fs.cell(row=row, column=mc).fill = _TOTAL_FILL
        fs.cell(row=row, column=mc).border = _THIN_BORDER

    c = fs.cell(row=row, column=5, value=f"Renamed: {renamed_count}")
    c.font = _ACTION_FONT_RENAMED
    c.fill = _TOTAL_FILL
    c.border = _THIN_BORDER
    c.alignment = _CENTER

    c = fs.cell(row=row, column=6, value=f"As-Is: {total - renamed_count}")
    c.font = _ACTION_FONT_ASIS
    c.fill = _TOTAL_FILL
    c.border = _THIN_BORDER
    c.alignment = _CENTER

    fs.row_dimensions[row].height = 26

    # ── Auto-width ───────────────────────────────────────────
    fs.column_dimensions["C"].width = max(max_orig_len  + 4, 20)
    fs.column_dimensions["D"].width = max(max_final_len + 4, 20)
    fs.column_dimensions["E"].width = 12  # Action
    fs.column_dimensions["F"].width = 14  # Result

    # ── Freeze & print ────────────────────────────────────────
    fs.freeze_panes = "B6"
    fs.print_options.horizontalCentered = True
    fs.page_margins.left = 0.5
    fs.page_margins.right = 0.5
