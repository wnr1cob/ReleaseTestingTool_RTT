"""
openpyxl style constants shared across the PDF Analyser report sheets.

All names are module-level constants; import them with a star import or
individually as needed:

    from src.core.pdf_analyzer.styles import _THIN_BORDER, _STATUS_FILLS
"""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# ── Borders ────────────────────────────────────────────────────
_NO_BORDER = Border()
_THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

# ── Fills ──────────────────────────────────────────────────────
_TITLE_FILL    = PatternFill(start_color="0D1B2A", end_color="0D1B2A", fill_type="solid")
_DARK_FILL     = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
_SECTION_FILL  = PatternFill(start_color="1A1A40", end_color="1A1A40", fill_type="solid")
_TABLE_HEADER_FILL = PatternFill(start_color="0F3460", end_color="0F3460", fill_type="solid")
_ROW_EVEN      = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
_ROW_ODD       = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_TOTAL_FILL    = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

_STATUS_FILLS = {
    "Passed":    PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid"),
    "Failed":    PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid"),
    "Error":     PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
    "Undefined": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),
    "Unknown":   PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"),
}

_CARD_FILLS = {
    "Passed":    PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid"),
    "Failed":    PatternFill(start_color="B71C1C", end_color="B71C1C", fill_type="solid"),
    "Error":     PatternFill(start_color="E65100", end_color="E65100", fill_type="solid"),
    "Undefined": PatternFill(start_color="37474F", end_color="37474F", fill_type="solid"),
    "Total":     PatternFill(start_color="01579B", end_color="01579B", fill_type="solid"),
}

# ── Fonts ──────────────────────────────────────────────────────
_TITLE_FONT        = Font(name="Segoe UI", size=18, bold=True,  color="00D4FF")
_SUBTITLE_FONT     = Font(name="Segoe UI", size=10,             color="8892B0")
_CARD_LABEL_FONT   = Font(name="Segoe UI", size=9,              color="AAAAAA")
_CARD_VALUE_FONT   = Font(name="Segoe UI", size=16, bold=True,  color="FFFFFF")
_CARD_PCT_FONT     = Font(name="Segoe UI", size=9,              color="CCCCCC")
_TABLE_HEADER_FONT = Font(name="Segoe UI", size=11, bold=True,  color="FFFFFF")
_BODY_FONT         = Font(name="Segoe UI", size=10,             color="333333")
_BODY_BOLD         = Font(name="Segoe UI", size=10, bold=True,  color="333333")

_STATUS_FONTS = {
    "Passed":    Font(name="Segoe UI", size=10, bold=True, color="1B5E20"),
    "Failed":    Font(name="Segoe UI", size=10, bold=True, color="B71C1C"),
    "Error":     Font(name="Segoe UI", size=10, bold=True, color="E65100"),
    "Undefined": Font(name="Segoe UI", size=10, bold=True, color="616161"),
    "Unknown":   Font(name="Segoe UI", size=10, bold=True, color="9E9E9E"),
}

# ── Alignment ──────────────────────────────────────────────────
_CENTER       = Alignment(horizontal="center",  vertical="center")
_LEFT_CENTER  = Alignment(horizontal="left",    vertical="center")
_RIGHT_CENTER = Alignment(horizontal="right",   vertical="center")
