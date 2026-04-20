"""
SW/Variant Analyzer page – analyze SW versions and variant information from PDF files.

Features:
  • Select check type: SW Check, Variant Check, or Both
  • Browse and select a folder to scan recursively for PDFs
  • Extract SW and Variant data from all PDFs in the folder
  • Display found versions as checkboxes for user selection
  • Generate Excel report of unselected versions with their PDF files
"""
import logging
import os
import threading
import time
from tkinter import filedialog, messagebox
from pathlib import Path

import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.gui.styles.theme import AppTheme as T
from src.gui.widgets.hover_button import RttButton
from src.gui.widgets.segmented_progress import SegmentedProgressBar
from src.utils import fmt_elapsed as _fmt_elapsed
from src.utils.config_manager import get_config_dir as _get_config_dir
from src.core.systemtestliste.utils import (
    load_variant_map,
    extract_sw_name,
    extract_variant_from_swfl,
)


logger = logging.getLogger(__name__)


class SWVariantAnalyzerPage(ctk.CTkFrame):
    """SW/Variant Analyzer page."""

    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)

        self._selected_dir: str = ""
        self._check_type: str = "both"  # 'sw', 'variant', or 'both'
        self._pdf_data: dict[str, dict] = {}  # {filename: {'sw': set, 'variant': set}}
        self._sw_versions: set[str] = set()
        self._variant_versions: set[str] = set()
        self._selected_sw: set[str] = set()
        self._selected_variant: set[str] = set()

        # Threading state
        self._poll_lock = threading.Lock()
        self._pending_status: tuple | None = None
        self._pending_message: str = ""
        self._pending_progress: float = 0.0
        self._poll_running: bool = False
        self._cancel_event = threading.Event()
        self._worker_thread: threading.Thread | None = None

        # variant map
        self._variant_map = self._load_variant_map()

        self._build()

    def _load_variant_map(self) -> dict[str, str]:
        """Load variant mapping from Variant_Info.txt."""
        try:
            config_dir = _get_config_dir()
            variant_path = os.path.join(config_dir, "Variant_Info.txt")
            return load_variant_map(path=variant_path)
        except Exception as e:
            logger.warning(f"Failed to load variant map: {e}")
            return {}

    def _build(self):
        """Build the page layout."""
        # Scrollable container so all cards are reachable
        self._scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self._scroll.pack(fill="both", expand=True)

        # ── Page title ──────────────────────────────────────────
        title_frame = ctk.CTkFrame(self._scroll, fg_color="transparent")
        title_frame.pack(fill="x", padx=30, pady=(25, 5))

        ctk.CTkLabel(
            title_frame,
            text="SW/Variant Analyzer",
            font=(T.FONT_FAMILY, T.FONT_SIZE_TITLE, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(side="left")

        ctk.CTkLabel(
            title_frame,
            text="Extract and analyze SW versions and variants from PDF files",
            font=(T.FONT_FAMILY, T.FONT_SIZE_SMALL),
            text_color=T.TEXT_SECONDARY,
        ).pack(side="left", padx=(15, 0), pady=(6, 0))

        # ── Check type card ─────────────────────────────────────
        type_card = ctk.CTkFrame(
            self._scroll,
            corner_radius=T.CARD_CORNER,
            fg_color=T.BG_CARD,
            border_width=1,
            border_color=T.BORDER_COLOR,
        )
        type_card.pack(fill="x", padx=30, pady=(20, 15))

        ctk.CTkLabel(
            type_card,
            text="Analysis Type",
            font=(T.FONT_FAMILY, T.FONT_SIZE_HEADING, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(anchor="w", padx=20, pady=(18, 12))

        types_frame = ctk.CTkFrame(type_card, fg_color="transparent")
        types_frame.pack(fill="x", padx=20, pady=(0, 20))

        self._check_var = ctk.StringVar(value="both")

        for text, value in [("SW Check", "sw"), ("Variant Check", "variant"), ("Both", "both")]:
            ctk.CTkRadioButton(
                types_frame,
                text=text,
                font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
                text_color=T.TEXT_PRIMARY,
                fg_color=T.ACCENT_PRIMARY,
                hover_color=T.SIDEBAR_BTN_HOVER,
                variable=self._check_var,
                value=value,
            ).pack(anchor="w", pady=4)

        # ── Directory selection card ────────────────────────────
        dir_card = ctk.CTkFrame(
            self._scroll,
            corner_radius=T.CARD_CORNER,
            fg_color=T.BG_CARD,
            border_width=1,
            border_color=T.BORDER_COLOR,
        )
        dir_card.pack(fill="x", padx=30, pady=(0, 15))

        ctk.CTkLabel(
            dir_card,
            text="Select Directory",
            font=(T.FONT_FAMILY, T.FONT_SIZE_HEADING, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(anchor="w", padx=20, pady=(18, 12))

        browse_row = ctk.CTkFrame(dir_card, fg_color="transparent")
        browse_row.pack(fill="x", padx=20, pady=(0, 20))

        self._path_entry = ctk.CTkEntry(
            browse_row,
            placeholder_text="No directory selected...",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            fg_color=T.BG_SIDEBAR,
            text_color=T.TEXT_PRIMARY,
            border_color=T.BORDER_COLOR,
            corner_radius=T.BUTTON_CORNER,
            height=38,
            state="disabled",
        )
        self._path_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        self._browse_btn = RttButton(
            browse_row,
            text="Browse",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            height=38,
            width=120,
            corner_radius=T.BUTTON_CORNER,
            fg_color=T.ACCENT_PRIMARY,
            hover_color=T.SIDEBAR_BTN_HOVER,
            text_color="#000000",
            command=self._browse_directory,
        )
        self._browse_btn.pack(side="right")

        # ── Status label ────────────────────────────────────────
        self._status_label = ctk.CTkLabel(
            self._scroll,
            text="Ready to analyze",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            text_color=T.TEXT_SECONDARY,
        )
        self._status_label.pack(padx=30, pady=(15, 0))

        # ── Progress bar ─────────────────────────────────────────
        self._progress_bar = SegmentedProgressBar(
            self._scroll,
            segments=[
                {"label": "Analyzing", "color": T.ACCENT_PRIMARY},
            ],
        )
        self._progress_bar.pack(fill="x", padx=30, pady=(8, 0))

        # ── Results container ───────────────────────────────────
        self._results_card = ctk.CTkFrame(
            self._scroll,
            corner_radius=T.CARD_CORNER,
            fg_color=T.BG_CARD,
            border_width=1,
            border_color=T.BORDER_COLOR,
        )
        self._results_card.pack(fill="both", expand=True, padx=30, pady=(15, 30))

        ctk.CTkLabel(
            self._results_card,
            text="Found Versions",
            font=(T.FONT_FAMILY, T.FONT_SIZE_HEADING, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(anchor="w", padx=20, pady=(18, 12))

        # Inner frame for checkboxes
        self._versions_frame = ctk.CTkFrame(self._results_card, fg_color="transparent")
        self._versions_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self._versions_label = ctk.CTkLabel(
            self._versions_frame,
            text="No analysis performed yet",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            text_color=T.TEXT_SECONDARY,
        )
        self._versions_label.pack(padx=10, pady=10)

        # ── Action buttons ──────────────────────────────────────
        btn_frame = ctk.CTkFrame(self._scroll, fg_color="transparent")
        btn_frame.pack(fill="x", padx=30, pady=(0, 30))

        self._analyze_btn = RttButton(
            btn_frame,
            text="Analyze",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            height=38,
            corner_radius=T.BUTTON_CORNER,
            fg_color=T.ACCENT_PRIMARY,
            hover_color=T.SIDEBAR_BTN_HOVER,
            text_color="#000000",
            command=self._start_analysis,
        )
        self._analyze_btn.pack(side="left", padx=(0, 10))

        self._export_btn = RttButton(
            btn_frame,
            text="Export Unselected",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            height=38,
            corner_radius=T.BUTTON_CORNER,
            fg_color=T.ACCENT_SECONDARY,
            hover_color=T.SIDEBAR_BTN_HOVER,
            text_color=T.TEXT_BRIGHT,
            command=self._export_unselected,
            state="disabled",
        )
        self._export_btn.pack(side="left", padx=(0, 10))

        self._clear_btn = RttButton(
            btn_frame,
            text="Clear",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            height=38,
            corner_radius=T.BUTTON_CORNER,
            fg_color=T.BORDER_COLOR,
            hover_color=T.SIDEBAR_BTN_HOVER,
            text_color=T.TEXT_PRIMARY,
            command=self._clear_analysis,
            state="disabled",
        )
        self._clear_btn.pack(side="left")

        # Start polling for updates
        self._poll_running = True
        self._poll_updates()

    def _browse_directory(self):
        """Open directory browser dialog."""
        folder = filedialog.askdirectory(title="Select Folder to Analyze")
        if folder:
            self._selected_dir = folder
            self._path_entry.configure(state="normal")
            self._path_entry.delete(0, "end")
            self._path_entry.insert(0, folder)
            self._path_entry.configure(state="disabled")

    def _start_analysis(self):
        """Start PDF analysis in background thread."""
        if not self._selected_dir:
            messagebox.showwarning("No Directory", "Please select a directory first")
            return

        self._analyze_btn.configure(state="disabled")
        self._export_btn.configure(state="disabled")
        self._clear_btn.configure(state="disabled")
        self._browse_btn.configure(state="disabled")

        self._pdf_data.clear()
        self._sw_versions.clear()
        self._variant_versions.clear()
        self._selected_sw.clear()
        self._selected_variant.clear()
        self._cancel_event.clear()

        # Reset progress bar
        self._progress_bar.reset()

        # Start worker thread
        self._worker_thread = threading.Thread(target=self._analyze_worker, daemon=True)
        self._worker_thread.start()

    def _analyze_worker(self):
        """Worker thread: scan folder and extract SW/Variant data."""
        try:
            check_type = self._check_var.get()
            start_time = time.monotonic()

            # Scan for PDFs
            self._set_status("Scanning for PDF files...", T.TEXT_SECONDARY)
            pdf_files = self._find_pdfs(self._selected_dir)

            if not pdf_files:
                self._set_status("No PDF files found in directory", T.ACCENT_DANGER)
                return

            self._set_status(f"Extracting data from {len(pdf_files)} PDF files...", T.TEXT_SECONDARY)

            # Process each PDF
            for i, pdf_path in enumerate(pdf_files):
                if self._cancel_event.is_set():
                    self._set_status("Analysis cancelled", T.TEXT_SECONDARY)
                    return

                filename = os.path.basename(pdf_path)
                try:
                    sw_versions, variant_versions = self._extract_versions(pdf_path, check_type)
                    self._pdf_data[filename] = {
                        "path": pdf_path,
                        "sw": sw_versions,
                        "variant": variant_versions,
                    }

                    # Collect all versions
                    self._sw_versions.update(sw_versions)
                    self._variant_versions.update(variant_versions)

                except Exception as e:
                    logger.warning(f"Failed to process {filename}: {e}")

                # Update progress bar
                progress_value = (i + 1) / len(pdf_files)
                self._set_progress(progress_value)

                progress = f"Processed {i + 1}/{len(pdf_files)}"
                elapsed = time.monotonic() - start_time
                self._set_status(f"{progress} ({_fmt_elapsed(elapsed)})", T.TEXT_SECONDARY)

            elapsed = time.monotonic() - start_time
            self._set_progress(1.0)  # Set to complete
            self._set_status(
                f"Analysis complete: {len(self._pdf_data)} files processed ({_fmt_elapsed(elapsed)})",
                T.ACCENT_SUCCESS,
            )

            # Render results UI
            self._render_versions_ui()

        except Exception as e:
            logger.exception(f"Analysis failed: {e}")
            self._set_status(f"Error: {e}", T.ACCENT_DANGER)

        finally:
            self._analyze_btn.configure(state="normal")
            self._browse_btn.configure(state="normal")

    def _find_pdfs(self, folder: str) -> list[str]:
        """Recursively find all PDF files in folder."""
        pdfs = []
        for root, _dirs, files in os.walk(folder):
            for f in files:
                if f.lower().endswith(".pdf"):
                    pdfs.append(os.path.join(root, f))
        return sorted(pdfs)

    def _extract_versions(self, pdf_path: str, check_type: str) -> tuple[set[str], set[str]]:
        """Extract SW and variant versions from a PDF."""
        sw_set = set()
        variant_set = set()

        try:
            import pdfplumber

            with pdfplumber.open(pdf_path) as pdf:
                # Try pages 0-4 for extraction
                text_pages = []
                for idx in range(min(5, len(pdf.pages))):
                    page_text = pdf.pages[idx].extract_text() or ""
                    text_pages.append(page_text)

            combined_text = "\n".join(text_pages)

            # Extract SW versions
            if check_type in ("sw", "both"):
                sw = extract_sw_name(combined_text)
                if sw:
                    sw_set.add(sw)

            # Extract variant versions
            if check_type in ("variant", "both"):
                variant = extract_variant_from_swfl(combined_text, self._variant_map)
                if variant:
                    variant_set.add(variant)

        except Exception as e:
            logger.debug(f"PDF extraction error for {pdf_path}: {e}")

        return sw_set, variant_set

    def _render_versions_ui(self):
        """Render checkboxes for found versions."""
        # Clear old widgets
        for widget in self._versions_frame.winfo_children():
            widget.destroy()

        check_type = self._check_var.get()
        has_content = False

        # SW checkboxes
        if check_type in ("sw", "both") and self._sw_versions:
            ctk.CTkLabel(
                self._versions_frame,
                text="SW Versions:",
                font=(T.FONT_FAMILY, T.FONT_SIZE_BODY, "bold"),
                text_color=T.TEXT_BRIGHT,
            ).pack(anchor="w", padx=10, pady=(10, 5))

            for sw in sorted(self._sw_versions):
                var = ctk.BooleanVar(value=True)
                cb = ctk.CTkCheckBox(
                    self._versions_frame,
                    text=sw,
                    font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
                    text_color=T.TEXT_PRIMARY,
                    fg_color=T.ACCENT_PRIMARY,
                    variable=var,
                    command=lambda s=sw, v=var: self._update_selection("sw", s, v),
                )
                cb.pack(anchor="w", padx=20, pady=2)
                self._selected_sw.add(sw)

            has_content = True

        # Variant checkboxes
        if check_type in ("variant", "both") and self._variant_versions:
            if has_content:
                ctk.CTkLabel(
                    self._versions_frame,
                    text="",
                    font=(T.FONT_FAMILY, 8),
                ).pack()

            ctk.CTkLabel(
                self._versions_frame,
                text="Variant Versions:",
                font=(T.FONT_FAMILY, T.FONT_SIZE_BODY, "bold"),
                text_color=T.TEXT_BRIGHT,
            ).pack(anchor="w", padx=10, pady=(10, 5))

            for variant in sorted(self._variant_versions):
                var = ctk.BooleanVar(value=True)
                cb = ctk.CTkCheckBox(
                    self._versions_frame,
                    text=variant,
                    font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
                    text_color=T.TEXT_PRIMARY,
                    fg_color=T.ACCENT_PRIMARY,
                    variable=var,
                    command=lambda v=variant, var=var: self._update_selection("variant", v, var),
                )
                cb.pack(anchor="w", padx=20, pady=2)
                self._selected_variant.add(variant)

            has_content = True

        if not has_content:
            self._versions_label.configure(
                text="No versions found matching the selected analysis type"
            )
            self._versions_label.pack(padx=10, pady=10)

        self._export_btn.configure(state="normal")
        self._clear_btn.configure(state="normal")

    def _update_selection(self, check_type: str, version: str, var: ctk.BooleanVar):
        """Update selection when checkbox is toggled."""
        if check_type == "sw":
            if var.get():
                self._selected_sw.add(version)
            else:
                self._selected_sw.discard(version)
        else:
            if var.get():
                self._selected_variant.add(version)
            else:
                self._selected_variant.discard(version)

    def _export_unselected(self):
        """Export unselected versions to Excel."""
        try:
            check_type = self._check_var.get()
            unselected_data = self._collect_unselected_data(check_type)

            # Check if there's any data to export
            has_sw_data = bool(unselected_data.get('sw'))
            has_variant_data = bool(unselected_data.get('variant'))
            
            if not has_sw_data and not has_variant_data:
                messagebox.showinfo("No Data", "All versions are selected. Nothing to export.")
                return

            output_dir = os.path.join(os.path.expanduser("~"), "Downloads")
            excel_path = self._create_excel_report(unselected_data, output_dir)

            messagebox.showinfo("Export Complete", f"Report saved to:\n{excel_path}")
            os.startfile(excel_path)

        except Exception as e:
            logger.exception(f"Export failed: {e}")
            messagebox.showerror("Export Failed", str(e))

    def _collect_unselected_data(self, check_type: str) -> dict[str, dict]:
        """Collect PDFs for unselected versions.
        
        Returns dict with structure:
        {
            'sw': {sw_version: [pdf_names]},
            'variant': {variant_version: [pdf_names]}
        }
        """
        data = {
            'sw': {},
            'variant': {},
        }

        for filename, pdf_info in self._pdf_data.items():
            sw_versions = pdf_info.get("sw", set())
            variant_versions = pdf_info.get("variant", set())

            # Add unselected SW versions
            if check_type in ("sw", "both"):
                unselected_sw = sw_versions - self._selected_sw
                for sw in unselected_sw:
                    if sw not in data['sw']:
                        data['sw'][sw] = []
                    data['sw'][sw].append(filename)

            # Add unselected variant versions
            if check_type in ("variant", "both"):
                unselected_variant = variant_versions - self._selected_variant
                for variant in unselected_variant:
                    if variant not in data['variant']:
                        data['variant'][variant] = []
                    data['variant'][variant].append(filename)

        return data

    def _create_excel_report(self, data: dict[str, dict], output_dir: str) -> str:
        """Create Excel report with SW and Variant mismatch tabs."""
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        check_type = self._check_var.get()
        
        # ── SW Mismatch Sheet ──────────────────────────────────
        if check_type in ("sw", "both") and data.get('sw'):
            ws_sw = wb.create_sheet("SW mismatch")
            self._populate_sw_sheet(ws_sw, data['sw'])
        
        # ── Variant Mismatch Sheet ─────────────────────────────
        if check_type in ("variant", "both") and data.get('variant'):
            ws_variant = wb.create_sheet("Variant mismatch")
            self._populate_variant_sheet(ws_variant, data['variant'])
        
        # If no sheets were created, add an empty info sheet
        if len(wb.sheetnames) == 0:
            ws_info = wb.create_sheet("Info")
            ws_info["A1"] = "No mismatches found. All versions are selected."
        
        # Save
        timestamp = os.getenv("RTT_TIMESTAMP", "")
        filename = f"SW_Variant_Mismatch_{timestamp or 'report'}.xlsx"
        output_path = os.path.join(output_dir, filename)
        wb.save(output_path)
        
        return output_path
    
    def _populate_sw_sheet(self, ws, sw_data: dict[str, list]):
        """Populate the SW mismatch sheet."""
        # Header
        header_fill = PatternFill(start_color="00D4FF", end_color="00D4FF", fill_type="solid")
        header_font = Font(name="Segoe UI", size=12, bold=True, color="000000")
        header_alignment = Alignment(horizontal="left", vertical="center")
        
        ws["A1"] = "SW Version"
        ws["B1"] = "PDF Name"
        
        for cell in ["A1", "B1"]:
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = header_alignment
        
        # Data rows
        row = 2
        for sw_version in sorted(sw_data.keys()):
            for pdf_name in sorted(sw_data[sw_version]):
                # Remove .pdf extension
                pdf_name_no_ext = os.path.splitext(pdf_name)[0]
                ws[f"A{row}"] = sw_version
                ws[f"B{row}"] = pdf_name_no_ext
                row += 1
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 70
    
    def _populate_variant_sheet(self, ws, variant_data: dict[str, list]):
        """Populate the Variant mismatch sheet."""
        # Header
        header_fill = PatternFill(start_color="A371F7", end_color="A371F7", fill_type="solid")
        header_font = Font(name="Segoe UI", size=12, bold=True, color="000000")
        header_alignment = Alignment(horizontal="left", vertical="center")
        
        ws["A1"] = "Variant Version"
        ws["B1"] = "PDF Name"
        
        for cell in ["A1", "B1"]:
            ws[cell].fill = header_fill
            ws[cell].font = header_font
            ws[cell].alignment = header_alignment
        
        # Data rows
        row = 2
        for variant_version in sorted(variant_data.keys()):
            for pdf_name in sorted(variant_data[variant_version]):
                # Remove .pdf extension
                pdf_name_no_ext = os.path.splitext(pdf_name)[0]
                ws[f"A{row}"] = variant_version
                ws[f"B{row}"] = pdf_name_no_ext
                row += 1
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 70


    def _clear_analysis(self):
        """Clear analysis results."""
        self._pdf_data.clear()
        self._sw_versions.clear()
        self._variant_versions.clear()
        self._selected_sw.clear()
        self._selected_variant.clear()

        # Clear UI
        for widget in self._versions_frame.winfo_children():
            widget.destroy()

        # Recreate the placeholder label
        self._versions_label = ctk.CTkLabel(
            self._versions_frame,
            text="No analysis performed yet",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            text_color=T.TEXT_SECONDARY,
        )
        self._versions_label.pack(padx=10, pady=10)

        self._export_btn.configure(state="disabled")
        self._clear_btn.configure(state="disabled")
        self._status_label.configure(text="Ready to analyze", text_color=T.TEXT_SECONDARY)
        self._progress_bar.reset()

    def _set_status(self, text: str, color: str):
        """Thread-safe status update."""
        with self._poll_lock:
            self._pending_message = text
            self._pending_status = (text, color)

    def _set_progress(self, value: float):
        """Thread-safe progress bar update."""
        with self._poll_lock:
            self._pending_progress = value

    def _poll_updates(self):
        """Poll for UI updates from worker thread."""
        if self._poll_running:
            with self._poll_lock:
                if self._pending_status:
                    text, color = self._pending_status
                    self._status_label.configure(text=text, text_color=color)
                    self._pending_status = None

                if self._pending_progress > 0.0:
                    self._progress_bar.set_segment(0, self._pending_progress)

        if self._poll_running:
            self.after(100, self._poll_updates)

    def pack_forget(self):
        """Called when page is hidden."""
        self._cancel_event.set()
        super().pack_forget()
