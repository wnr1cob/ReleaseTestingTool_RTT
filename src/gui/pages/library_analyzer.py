"""
Library Analyzer page – analyze library versions from PDF files.

Features:
  • Browse and select a folder to scan recursively for PDFs
  • Extract library version data from all PDFs in the folder
  • Identify PDFs without a library version
  • Export missing versions to Excel
"""
import logging
import os
import threading
import time
from tkinter import filedialog, messagebox

import customtkinter as ctk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from src.gui.styles.theme import AppTheme as T
from src.gui.widgets.hover_button import RttButton
from src.gui.widgets.segmented_progress import SegmentedProgressBar
from src.utils import fmt_elapsed as _fmt_elapsed
from src.core.systemtestliste.utils import extract_library_version


logger = logging.getLogger(__name__)


class LibraryAnalyzerPage(ctk.CTkFrame):
    """Library Analyzer page."""

    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)

        self._selected_dir: str = ""
        self._pdf_data: dict[str, dict] = {}  # {filename: {'has_version': bool, 'version': str}}
        self._missing_version_pdfs: list[str] = []

        # Threading state
        self._poll_lock = threading.Lock()
        self._pending_status: tuple | None = None
        self._pending_message: str = ""
        self._pending_progress: float = 0.0
        self._poll_running: bool = False
        self._cancel_event = threading.Event()
        self._worker_thread: threading.Thread | None = None

        self._build()

    def _build(self):
        """Build the page layout."""
        # Scrollable container so all cards are reachable
        self._scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self._scroll.pack(fill="both", expand=True)

        # ── Page title ──────────────────────────────────────────
        title_frame = ctk.CTkFrame(self._scroll, fg_color="transparent")
        title_frame.pack(fill="x", padx=30, pady=(25, 5))

        ctk.CTkLabel(
            title_frame,
            text="Library Analyzer",
            font=(T.FONT_FAMILY, T.FONT_SIZE_TITLE, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(side="left")

        ctk.CTkLabel(
            title_frame,
            text="Extract and verify library versions from PDF files",
            font=(T.FONT_FAMILY, T.FONT_SIZE_SMALL),
            text_color=T.TEXT_SECONDARY,
        ).pack(side="left", padx=(15, 0), pady=(6, 0))

        # ── Directory selection card ────────────────────────────
        dir_card = ctk.CTkFrame(
            self._scroll,
            corner_radius=T.CARD_CORNER,
            fg_color=T.BG_CARD,
            border_width=1,
            border_color=T.BORDER_COLOR,
        )
        dir_card.pack(fill="x", padx=30, pady=(20, 15))

        ctk.CTkLabel(
            dir_card,
            text="Select Directory",
            font=(T.FONT_FAMILY, T.FONT_SIZE_HEADING, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(anchor="w", padx=20, pady=(18, 12))

        browse_row = ctk.CTkFrame(dir_card, fg_color="transparent")
        browse_row.pack(fill="x", padx=20, pady=(0, 20))

        self._path_entry = ctk.CTkEntry(
            browse_row,
            placeholder_text="No directory selected...",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            fg_color=T.BG_SIDEBAR,
            text_color=T.TEXT_PRIMARY,
            border_color=T.BORDER_COLOR,
            corner_radius=T.BUTTON_CORNER,
            height=38,
            state="disabled",
        )
        self._path_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        self._browse_btn = RttButton(
            browse_row,
            text="Browse",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            height=38,
            width=120,
            corner_radius=T.BUTTON_CORNER,
            fg_color=T.ACCENT_PRIMARY,
            hover_color=T.SIDEBAR_BTN_HOVER,
            text_color="#000000",
            command=self._browse_directory,
        )
        self._browse_btn.pack(side="right")

        # ── Status label ────────────────────────────────────────
        self._status_label = ctk.CTkLabel(
            self._scroll,
            text="Ready to analyze",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            text_color=T.TEXT_SECONDARY,
        )
        self._status_label.pack(padx=30, pady=(15, 0))

        # ── Progress bar ─────────────────────────────────────────
        self._progress_bar = SegmentedProgressBar(
            self._scroll,
            segments=[
                {"label": "Analyzing", "color": T.ACCENT_PRIMARY},
            ],
        )
        self._progress_bar.pack(fill="x", padx=30, pady=(8, 0))

        # ── Results container ───────────────────────────────────
        self._results_card = ctk.CTkFrame(
            self._scroll,
            corner_radius=T.CARD_CORNER,
            fg_color=T.BG_CARD,
            border_width=1,
            border_color=T.BORDER_COLOR,
        )
        self._results_card.pack(fill="both", expand=True, padx=30, pady=(15, 30))

        ctk.CTkLabel(
            self._results_card,
            text="Analysis Results",
            font=(T.FONT_FAMILY, T.FONT_SIZE_HEADING, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(anchor="w", padx=20, pady=(18, 12))

        # Inner frame for results
        self._results_frame = ctk.CTkFrame(self._results_card, fg_color="transparent")
        self._results_frame.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        self._results_label = ctk.CTkLabel(
            self._results_frame,
            text="No analysis performed yet",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            text_color=T.TEXT_SECONDARY,
        )
        self._results_label.pack(padx=10, pady=10)

        # ── Action buttons ──────────────────────────────────────
        btn_frame = ctk.CTkFrame(self._scroll, fg_color="transparent")
        btn_frame.pack(fill="x", padx=30, pady=(0, 30))

        self._analyze_btn = RttButton(
            btn_frame,
            text="Analyze",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            height=38,
            corner_radius=T.BUTTON_CORNER,
            fg_color=T.ACCENT_PRIMARY,
            hover_color=T.SIDEBAR_BTN_HOVER,
            text_color="#000000",
            command=self._start_analysis,
        )
        self._analyze_btn.pack(side="left", padx=(0, 10))

        self._export_btn = RttButton(
            btn_frame,
            text="Export Missing Versions",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            height=38,
            corner_radius=T.BUTTON_CORNER,
            fg_color=T.ACCENT_SECONDARY,
            hover_color=T.SIDEBAR_BTN_HOVER,
            text_color=T.TEXT_BRIGHT,
            command=self._export_missing,
            state="disabled",
        )
        self._export_btn.pack(side="left", padx=(0, 10))

        self._clear_btn = RttButton(
            btn_frame,
            text="Clear",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            height=38,
            corner_radius=T.BUTTON_CORNER,
            fg_color=T.BORDER_COLOR,
            hover_color=T.SIDEBAR_BTN_HOVER,
            text_color=T.TEXT_PRIMARY,
            command=self._clear_analysis,
            state="disabled",
        )
        self._clear_btn.pack(side="left")

        # Start polling for updates
        self._poll_running = True
        self._poll_updates()

    def _browse_directory(self):
        """Open directory browser dialog."""
        folder = filedialog.askdirectory(title="Select Folder to Analyze")
        if folder:
            self._selected_dir = folder
            self._path_entry.configure(state="normal")
            self._path_entry.delete(0, "end")
            self._path_entry.insert(0, folder)
            self._path_entry.configure(state="disabled")

    def _start_analysis(self):
        """Start PDF analysis in background thread."""
        if not self._selected_dir:
            messagebox.showwarning("No Directory", "Please select a directory first")
            return

        self._analyze_btn.configure(state="disabled")
        self._export_btn.configure(state="disabled")
        self._clear_btn.configure(state="disabled")
        self._browse_btn.configure(state="disabled")

        self._pdf_data.clear()
        self._missing_version_pdfs.clear()
        self._cancel_event.clear()

        # Reset progress bar
        self._progress_bar.reset()

        # Start worker thread
        self._worker_thread = threading.Thread(target=self._analyze_worker, daemon=True)
        self._worker_thread.start()

    def _analyze_worker(self):
        """Worker thread: scan folder and extract library version data."""
        try:
            start_time = time.monotonic()

            # Scan for PDFs
            self._set_status("Scanning for PDF files...", T.TEXT_SECONDARY)
            pdf_files = self._find_pdfs(self._selected_dir)

            if not pdf_files:
                self._set_status("No PDF files found in directory", T.ACCENT_DANGER)
                return

            self._set_status(f"Extracting library versions from {len(pdf_files)} PDF files...", T.TEXT_SECONDARY)

            # Process each PDF
            for i, pdf_path in enumerate(pdf_files):
                if self._cancel_event.is_set():
                    self._set_status("Analysis cancelled", T.TEXT_SECONDARY)
                    return

                filename = os.path.basename(pdf_path)
                try:
                    has_version, version = self._extract_library_version(pdf_path)
                    self._pdf_data[filename] = {
                        "path": pdf_path,
                        "has_version": has_version,
                        "version": version,
                    }

                    if not has_version:
                        self._missing_version_pdfs.append(filename)

                except Exception as e:
                    logger.warning(f"Failed to process {filename}: {e}")

                # Update progress bar
                progress_value = (i + 1) / len(pdf_files)
                self._set_progress(progress_value)

                progress = f"Processed {i + 1}/{len(pdf_files)}"
                elapsed = time.monotonic() - start_time
                self._set_status(f"{progress} ({_fmt_elapsed(elapsed)})", T.TEXT_SECONDARY)

            elapsed = time.monotonic() - start_time
            missing_count = len(self._missing_version_pdfs)
            self._set_progress(1.0)  # Set to complete
            self._set_status(
                f"Analysis complete: {len(self._pdf_data)} files processed, {missing_count} missing versions ({_fmt_elapsed(elapsed)})",
                T.ACCENT_SUCCESS,
            )

            # Render results UI
            self._render_results_ui()

        except Exception as e:
            logger.exception(f"Analysis failed: {e}")
            self._set_status(f"Error: {e}", T.ACCENT_DANGER)

        finally:
            self._analyze_btn.configure(state="normal")
            self._browse_btn.configure(state="normal")

    def _find_pdfs(self, folder: str) -> list[str]:
        """Recursively find all PDF files in folder."""
        pdfs = []
        for root, _dirs, files in os.walk(folder):
            for f in files:
                if f.lower().endswith(".pdf"):
                    pdfs.append(os.path.join(root, f))
        return sorted(pdfs)

    def _extract_library_version(self, pdf_path: str) -> tuple[bool, str]:
        """Extract library version from a PDF."""
        try:
            import pdfplumber

            with pdfplumber.open(pdf_path) as pdf:
                # Try pages 0-4 for extraction
                text_pages = []
                for idx in range(min(5, len(pdf.pages))):
                    page_text = pdf.pages[idx].extract_text() or ""
                    text_pages.append(page_text)

            combined_text = "\n".join(text_pages)

            # Extract library version using existing logic
            version = extract_library_version(combined_text)
            has_version = version is not None and version.strip() != ""

            return has_version, version or ""

        except Exception as e:
            logger.debug(f"PDF extraction error for {pdf_path}: {e}")
            return False, ""

    def _render_results_ui(self):
        """Render results summary."""
        # Clear old widgets
        for widget in self._results_frame.winfo_children():
            widget.destroy()

        total = len(self._pdf_data)
        missing = len(self._missing_version_pdfs)
        found = total - missing

        # Summary text
        summary_text = f"Total PDFs: {total} | With Version: {found} | Missing Version: {missing}"
        ctk.CTkLabel(
            self._results_frame,
            text=summary_text,
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            text_color=T.TEXT_PRIMARY,
        ).pack(anchor="w", padx=10, pady=10)

        # List missing versions if any
        if self._missing_version_pdfs:
            ctk.CTkLabel(
                self._results_frame,
                text="PDFs Missing Library Version:",
                font=(T.FONT_FAMILY, T.FONT_SIZE_BODY, "bold"),
                text_color=T.ACCENT_WARNING,
            ).pack(anchor="w", padx=10, pady=(10, 5))

            for pdf_name in sorted(self._missing_version_pdfs)[:10]:  # Show first 10
                pdf_name_no_ext = os.path.splitext(pdf_name)[0]
                ctk.CTkLabel(
                    self._results_frame,
                    text=f"  • {pdf_name_no_ext}",
                    font=(T.FONT_FAMILY, T.FONT_SIZE_SMALL),
                    text_color=T.TEXT_SECONDARY,
                ).pack(anchor="w", padx=20, pady=1)

            if len(self._missing_version_pdfs) > 10:
                ctk.CTkLabel(
                    self._results_frame,
                    text=f"  ... and {len(self._missing_version_pdfs) - 10} more",
                    font=(T.FONT_FAMILY, T.FONT_SIZE_SMALL),
                    text_color=T.TEXT_SECONDARY,
                ).pack(anchor="w", padx=20, pady=1)

        self._export_btn.configure(state="normal")
        self._clear_btn.configure(state="normal")

    def _export_missing(self):
        """Export PDFs with missing library versions to Excel."""
        if not self._missing_version_pdfs:
            messagebox.showinfo("No Data", "All PDFs have library versions. Nothing to export.")
            return

        try:
            output_dir = os.path.join(os.path.expanduser("~"), "Downloads")
            excel_path = self._create_excel_report(output_dir)

            messagebox.showinfo("Export Complete", f"Report saved to:\n{excel_path}")
            os.startfile(excel_path)

        except Exception as e:
            logger.exception(f"Export failed: {e}")
            messagebox.showerror("Export Failed", str(e))

    def _create_excel_report(self, output_dir: str) -> str:
        """Create Excel report of PDFs missing library versions."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Missing Library Versions"

        # Header
        header_fill = PatternFill(start_color="00D4FF", end_color="00D4FF", fill_type="solid")
        header_font = Font(name="Segoe UI", size=12, bold=True, color="000000")
        header_alignment = Alignment(horizontal="left", vertical="center")

        ws["A1"] = "PDF Name"

        ws["A1"].fill = header_fill
        ws["A1"].font = header_font
        ws["A1"].alignment = header_alignment

        # Data rows
        row = 2
        for pdf_name in sorted(self._missing_version_pdfs):
            # Remove .pdf extension
            pdf_name_no_ext = os.path.splitext(pdf_name)[0]
            ws[f"A{row}"] = pdf_name_no_ext
            row += 1

        ws.column_dimensions["A"].width = 80

        # Save
        timestamp = os.getenv("RTT_TIMESTAMP", "")
        filename = f"Library_Missing_Versions_{timestamp or 'report'}.xlsx"
        output_path = os.path.join(output_dir, filename)
        wb.save(output_path)

        return output_path

    def _clear_analysis(self):
        """Clear analysis results."""
        self._pdf_data.clear()
        self._missing_version_pdfs.clear()

        # Clear UI
        for widget in self._results_frame.winfo_children():
            widget.destroy()

        # Recreate the placeholder label
        self._results_label = ctk.CTkLabel(
            self._results_frame,
            text="No analysis performed yet",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            text_color=T.TEXT_SECONDARY,
        )
        self._results_label.pack(padx=10, pady=10)

        self._export_btn.configure(state="disabled")
        self._clear_btn.configure(state="disabled")
        self._status_label.configure(text="Ready to analyze", text_color=T.TEXT_SECONDARY)
        self._progress_bar.reset()

    def _set_status(self, text: str, color: str):
        """Thread-safe status update."""
        with self._poll_lock:
            self._pending_message = text
            self._pending_status = (text, color)

    def _set_progress(self, value: float):
        """Thread-safe progress bar update."""
        with self._poll_lock:
            self._pending_progress = value

    def _poll_updates(self):
        """Poll for UI updates from worker thread."""
        if self._poll_running:
            with self._poll_lock:
                if self._pending_status:
                    text, color = self._pending_status
                    self._status_label.configure(text=text, text_color=color)
                    self._pending_status = None

                if self._pending_progress > 0.0:
                    self._progress_bar.set_segment(0, self._pending_progress)

        if self._poll_running:
            self.after(100, self._poll_updates)

    def pack_forget(self):
        """Called when page is hidden."""
        self._cancel_event.set()
        super().pack_forget()
