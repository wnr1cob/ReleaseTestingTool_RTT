"""
SystemTestListe Analyzer page – cross-reference PDF results against an Excel SystemTestListe.
"""
import os
import re
import threading
from tkinter import filedialog
import customtkinter as ctk
from src.gui.styles.theme import AppTheme as T
from src.gui.widgets.segmented_progress import SegmentedProgressBar


# Variant suffix pattern at the end of a tab name: _v3, _V14, _NA0, _NA5, _G5x, _G4x, etc.
_VARIANT_RE = re.compile(r"(_(?:[vV]\d+|NA\d+|G\d+x))$")


def _parse_sw_variant(tab_name: str) -> tuple[str, str]:
    """Split a tab name into (sw_name, variant). Returns (tab_name, '') if no variant found."""
    m = _VARIANT_RE.search(tab_name)
    if m:
        variant = m.group(1).lstrip("_")
        sw = tab_name[: m.start()]
        return sw, variant
    return tab_name, ""


def _cell_to_str(v) -> str:
    """Convert any Excel cell value to a clean string.
    - None / empty  → ''
    - float that is a whole number (e.g. 42.0) → '42'
    - datetime / date → ISO-formatted string
    - everything else → str(v)
    """
    if v is None or v == "":
        return ""
    import datetime as _dt
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else str(v)
    if isinstance(v, (_dt.datetime, _dt.date)):
        return v.isoformat()
    return str(v).strip()


class SystemTestListePage(ctk.CTkFrame):
    """SystemTestListe Analyzer page – verify PDF result/variant/SW info against an Excel sheet."""

    def __init__(self, master, **kwargs):
        super().__init__(master, fg_color="transparent", **kwargs)
        self._excel_file: str = ""
        self._pdf_dir: str = ""
        self._all_tabs: list[str] = []
        self._selected_tab: str = ""
        self._sw_name: str = ""
        self._variant: str = ""
        self._result_path: str = ""
        self._build()

    # ────────────────────────────────────────────────────────────
    # Build layout
    # ────────────────────────────────────────────────────────────
    def _build(self):
        # Scrollable container so all cards are reachable
        self._scroll = ctk.CTkScrollableFrame(self, fg_color="transparent")
        self._scroll.pack(fill="both", expand=True)

        # ── Page title ──────────────────────────────────────────
        title_frame = ctk.CTkFrame(self._scroll, fg_color="transparent")
        title_frame.pack(fill="x", padx=30, pady=(25, 5))

        ctk.CTkLabel(
            title_frame,
            text="SystemTestListe Analyzer",
            font=(T.FONT_FAMILY, T.FONT_SIZE_TITLE, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(side="left")

        ctk.CTkLabel(
            title_frame,
            text="Verify PDF results against the SystemTestListe Excel sheet",
            font=(T.FONT_FAMILY, T.FONT_SIZE_SMALL),
            text_color=T.TEXT_SECONDARY,
        ).pack(side="left", padx=(15, 0), pady=(6, 0))

        # ── Cards 1 & 4 · Excel (50%) + Folder (50%) side by side ──
        self._file_row = ctk.CTkFrame(self._scroll, fg_color="transparent")
        self._file_row.pack(fill="x", padx=30, pady=(20, 15))
        self._file_row.columnconfigure(0, weight=1)
        self._file_row.columnconfigure(1, weight=1)

        self._build_excel_card()

        # ── Cards 2 & 3 · Tab list (90%) + SW/Variant (10%) side by side
        self._side_row = ctk.CTkFrame(self._scroll, fg_color="transparent")
        self._side_row.pack(fill="x", padx=30, pady=(0, 15))
        self._side_row.columnconfigure(0, weight=9)  # 90%
        self._side_row.columnconfigure(1, weight=1)  # 10%

        self._build_tab_card()
        self._build_sw_variant_card()

        # ── Card 4 · Folder selection (built into _file_row) ──
        self._build_folder_card()

        # ── Card 5 · Verification options ───────────────────────
        self._build_options_card()

        # ── Bottom bar ──────────────────────────────────────────
        self._build_bottom_bar()

    # ── Card 1 ─────────────────────────────────────────────────
    def _build_excel_card(self):
        card = ctk.CTkFrame(
            self._file_row,
            corner_radius=T.CARD_CORNER,
            fg_color=T.BG_CARD,
            border_width=1,
            border_color=T.BORDER_COLOR,
        )
        card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        ctk.CTkLabel(
            card,
            text="📗  SystemTestListe Excel File",
            font=(T.FONT_FAMILY, T.FONT_SIZE_HEADING, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(anchor="w", padx=20, pady=(18, 12))

        row = ctk.CTkFrame(card, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=(0, 20))

        self._excel_entry = ctk.CTkEntry(
            row,
            placeholder_text="No Excel file selected...",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            fg_color=T.BG_SIDEBAR,
            text_color=T.TEXT_PRIMARY,
            border_color=T.BORDER_COLOR,
            corner_radius=T.BUTTON_CORNER,
            height=38,
            state="disabled",
        )
        self._excel_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        ctk.CTkButton(
            row,
            text="📂  Browse",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            height=38,
            width=120,
            corner_radius=T.BUTTON_CORNER,
            fg_color=T.ACCENT_SUCCESS,
            hover_color=T.SIDEBAR_BTN_HOVER,
            text_color=T.BG_DARK,
            command=self._browse_excel,
        ).pack(side="right")

    # ── Card 2 ─────────────────────────────────────────────────
    def _build_tab_card(self):
        self._tab_card = ctk.CTkFrame(
            self._side_row,
            corner_radius=T.CARD_CORNER,
            fg_color=T.BG_CARD,
            border_width=1,
            border_color=T.BORDER_COLOR,
        )
        self._tab_card.grid(row=0, column=0, sticky="nsew", padx=(0, 8))

        ctk.CTkLabel(
            self._tab_card,
            text="📑  Select Sheet / Tab",
            font=(T.FONT_FAMILY, T.FONT_SIZE_HEADING, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(anchor="w", padx=20, pady=(9, 5))

        # Filter row
        filter_row = ctk.CTkFrame(self._tab_card, fg_color="transparent")
        filter_row.pack(fill="x", padx=20, pady=(0, 5))

        ctk.CTkLabel(
            filter_row,
            text="🔎  Filter:",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            text_color=T.TEXT_SECONDARY,
        ).pack(side="left", padx=(0, 8))

        self._filter_var = ctk.StringVar()
        self._filter_var.trace_add("write", lambda *_: self._apply_filter())

        self._filter_entry = ctk.CTkEntry(
            filter_row,
            textvariable=self._filter_var,
            placeholder_text="Type to filter tabs...",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            fg_color=T.BG_SIDEBAR,
            text_color=T.TEXT_PRIMARY,
            border_color=T.BORDER_COLOR,
            corner_radius=T.BUTTON_CORNER,
            height=34,
        )
        self._filter_entry.pack(side="left", fill="x", expand=True)

        ctk.CTkButton(
            filter_row,
            text="✖",
            width=34,
            height=34,
            corner_radius=T.BUTTON_CORNER,
            fg_color=T.BG_SIDEBAR,
            hover_color=T.SIDEBAR_BTN_HOVER,
            text_color=T.TEXT_SECONDARY,
            command=lambda: self._filter_var.set(""),
        ).pack(side="left", padx=(6, 0))

        # List of tab buttons (plain frame – outer scroll handles scrolling)
        self._tab_list_frame = ctk.CTkFrame(
            self._tab_card,
            fg_color=T.BG_SIDEBAR,
            corner_radius=T.BUTTON_CORNER,
        )
        
        self._tab_list_frame.pack(fill="x", padx=20, pady=(0, 9))

        self._tab_buttons: dict[str, ctk.CTkButton] = {}

    # ── Card 3 ─────────────────────────────────────────────────
    def _build_sw_variant_card(self):
        self._sw_card = ctk.CTkFrame(
            self._side_row,
            corner_radius=T.CARD_CORNER,
            fg_color=T.BG_CARD,
            border_width=1,
            border_color=T.BORDER_COLOR,
        )
        self._sw_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        ctk.CTkLabel(
            self._sw_card,
            text="🏷️  Detected SW & Variant",
            font=(T.FONT_FAMILY, T.FONT_SIZE_HEADING, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(anchor="w", padx=20, pady=(9, 6))

        info_row = ctk.CTkFrame(self._sw_card, fg_color="transparent")
        info_row.pack(fill="x", padx=20, pady=(0, 9))

        # SW name chip
        sw_box = ctk.CTkFrame(info_row, fg_color=T.BG_SIDEBAR, corner_radius=T.BUTTON_CORNER)
        sw_box.pack(side="left", padx=(0, 20))

        ctk.CTkLabel(
            sw_box,
            text="SW",
            font=(T.FONT_FAMILY, T.FONT_SIZE_SMALL),
            text_color=T.TEXT_SECONDARY,
        ).pack(anchor="w", padx=12, pady=(4, 0))

        self._sw_label = ctk.CTkLabel(
            sw_box,
            text="—",
            font=(T.FONT_FAMILY, T.FONT_SIZE_HEADING, "bold"),
            text_color=T.ACCENT_PRIMARY,
        )
        self._sw_label.pack(padx=12, pady=(0, 4))

        # Variant chip
        var_box = ctk.CTkFrame(info_row, fg_color=T.BG_SIDEBAR, corner_radius=T.BUTTON_CORNER)
        var_box.pack(side="left")

        ctk.CTkLabel(
            var_box,
            text="Variant",
            font=(T.FONT_FAMILY, T.FONT_SIZE_SMALL),
            text_color=T.TEXT_SECONDARY,
        ).pack(anchor="w", padx=12, pady=(4, 0))

        self._variant_label = ctk.CTkLabel(
            var_box,
            text="—",
            font=(T.FONT_FAMILY, T.FONT_SIZE_HEADING, "bold"),
            text_color=T.ACCENT_WARNING,
        )
        self._variant_label.pack(padx=12, pady=(0, 4))

    # ── Card 4 ─────────────────────────────────────────────────
    def _build_folder_card(self):
        self._folder_card = ctk.CTkFrame(
            self._file_row,
            corner_radius=T.CARD_CORNER,
            fg_color=T.BG_CARD,
            border_width=1,
            border_color=T.BORDER_COLOR,
        )
        self._folder_card.grid(row=0, column=1, sticky="nsew", padx=(8, 0))

        ctk.CTkLabel(
            self._folder_card,
            text="📂  PDF Reports Directory",
            font=(T.FONT_FAMILY, T.FONT_SIZE_HEADING, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(anchor="w", padx=20, pady=(18, 12))

        row = ctk.CTkFrame(self._folder_card, fg_color="transparent")
        row.pack(fill="x", padx=20, pady=(0, 20))

        self._dir_entry = ctk.CTkEntry(
            row,
            placeholder_text="No directory selected...",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            fg_color=T.BG_SIDEBAR,
            text_color=T.TEXT_PRIMARY,
            border_color=T.BORDER_COLOR,
            corner_radius=T.BUTTON_CORNER,
            height=38,
            state="disabled",
        )
        self._dir_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        ctk.CTkButton(
            row,
            text="📁  Browse",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            height=38,
            width=120,
            corner_radius=T.BUTTON_CORNER,
            fg_color=T.ACCENT_PRIMARY,
            hover_color=T.SIDEBAR_BTN_HOVER,
            text_color=T.BG_DARK,
            command=self._browse_directory,
        ).pack(side="right")

    # ── Card 5 ─────────────────────────────────────────────────
    def _build_options_card(self):
        opt_card = ctk.CTkFrame(
            self._scroll,
            corner_radius=T.CARD_CORNER,
            fg_color=T.BG_CARD,
            border_width=1,
            border_color=T.BORDER_COLOR,
        )
        opt_card.pack(fill="x", padx=30, pady=(0, 15))

        ctk.CTkLabel(
            opt_card,
            text="🔍  Verification Options",
            font=(T.FONT_FAMILY, T.FONT_SIZE_HEADING, "bold"),
            text_color=T.TEXT_BRIGHT,
        ).pack(anchor="w", padx=20, pady=(18, 12))

        checks_row = ctk.CTkFrame(opt_card, fg_color="transparent")
        checks_row.pack(anchor="w", padx=20, pady=(0, 18))

        self._chk_result_var  = ctk.BooleanVar(value=True)
        self._chk_sw_var      = ctk.BooleanVar(value=True)
        self._chk_variant_var = ctk.BooleanVar(value=True)

        for var, label in [
            (self._chk_result_var,  "Result Status"),
            (self._chk_sw_var,      "SW Version"),
            (self._chk_variant_var, "Variant Information"),
        ]:
            ctk.CTkCheckBox(
                checks_row,
                text=label,
                font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
                text_color=T.TEXT_PRIMARY,
                fg_color=T.ACCENT_PRIMARY,
                hover_color=T.SIDEBAR_BTN_HOVER,
                variable=var,
                onvalue=True,
                offvalue=False,
            ).pack(side="left", padx=(0, 30))

    # ── Bottom bar ──────────────────────────────────────────────
    def _build_bottom_bar(self):
        self._bottom_card = ctk.CTkFrame(
            self,
            corner_radius=T.CARD_CORNER,
            fg_color=T.BG_CARD,
            border_width=1,
            border_color=T.BORDER_COLOR,
        )
        self._bottom_card.pack(fill="x", side="bottom", padx=30, pady=(0, 20))

        self._status_label = ctk.CTkLabel(
            self._bottom_card,
            text="  Select an Excel file to begin.",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
            text_color=T.TEXT_SECONDARY,
            anchor="w",
        )
        self._status_label.pack(anchor="w", padx=20, pady=(14, 6))

        self._progress = SegmentedProgressBar(
            self._bottom_card,
            segments=[
                {"label": "Reading Excel",  "color": T.ACCENT_SUCCESS},
                {"label": "Scanning PDFs",  "color": T.ACCENT_PRIMARY},
                {"label": "Report",         "color": T.ACCENT_SECONDARY},
            ],
        )
        self._progress.pack(fill="x", padx=20, pady=(0, 10))

        btn_row = ctk.CTkFrame(self._bottom_card, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=(0, 14))

        self._start_btn = ctk.CTkButton(
            btn_row,
            text="▶  Start",
            font=(T.FONT_FAMILY, T.FONT_SIZE_BODY, "bold"),
            height=40,
            width=140,
            corner_radius=T.BUTTON_CORNER,
            fg_color=T.ACCENT_SUCCESS,
            hover_color="#00c853",
            text_color=T.BG_DARK,
            command=self._start_analysis,
        )
        self._start_btn.pack(side="right")

    # ────────────────────────────────────────────────────────────
    # Excel browser & tab loading
    # ────────────────────────────────────────────────────────────
    def _browse_excel(self):
        path = filedialog.askopenfilename(
            title="Select SystemTestListe Excel File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
        )
        if not path:
            return
        self._excel_file = path
        self._excel_entry.configure(state="normal")
        self._excel_entry.delete(0, "end")
        self._excel_entry.insert(0, path)
        self._excel_entry.configure(state="disabled")

        # Reset downstream state
        self._selected_tab = ""
        self._sw_name = ""
        self._variant = ""
        self._sw_label.configure(text="—")
        self._variant_label.configure(text="—")

        self._status_label.configure(
            text="  Reading Excel sheets…", text_color=T.TEXT_SECONDARY
        )
        threading.Thread(target=self._load_tabs, daemon=True).start()

    def _load_tabs(self):
        """Read sheet names from the Excel file (background thread)."""
        try:
            ext = os.path.splitext(self._excel_file)[1].lower()
            if ext == ".xls":
                import xlrd
                wb = xlrd.open_workbook(self._excel_file)
                tabs = wb.sheet_names()
            else:
                import openpyxl
                wb = openpyxl.load_workbook(self._excel_file, read_only=True, data_only=True)
                tabs = wb.sheetnames
                wb.close()
        except Exception as e:
            self._ui(lambda: self._status_label.configure(
                text=f"  Failed to read Excel: {e}",
                text_color=T.ACCENT_DANGER,
            ))
            return

        self._all_tabs = tabs
        self._ui(self._populate_tab_list)

    def _populate_tab_list(self):
        """Build tab buttons inside the scrollable list (main thread)."""
        # Clear previous buttons
        for btn in self._tab_buttons.values():
            btn.destroy()
        self._tab_buttons.clear()
        self._filter_var.set("")

        for tab in self._all_tabs:
            btn = ctk.CTkButton(
                self._tab_list_frame,
                text=tab,
                font=(T.FONT_FAMILY, T.FONT_SIZE_BODY),
                anchor="w",
                height=32,
                corner_radius=T.BUTTON_CORNER,
                fg_color="transparent",
                hover_color=T.SIDEBAR_BTN_ACTIVE_BG,
                text_color=T.TEXT_PRIMARY,
                border_width=0,
                command=lambda t=tab: self._on_tab_select(t),
            )
            btn.pack(fill="x", pady=2, padx=4)
            self._tab_buttons[tab] = btn

        self._status_label.configure(
            text=f"  {len(self._all_tabs)} sheet(s) loaded. Select a tab.",
            text_color=T.TEXT_SECONDARY,
        )

    # ── Filter ──────────────────────────────────────────────────
    def _apply_filter(self):
        query = self._filter_var.get().lower()
        for tab, btn in self._tab_buttons.items():
            if query in tab.lower():
                btn.pack(fill="x", pady=2, padx=4)
            else:
                btn.pack_forget()

    # ── Tab selection ───────────────────────────────────────────
    def _on_tab_select(self, tab: str):
        self._selected_tab = tab

        # Highlight selected button, reset all others
        for name, btn in self._tab_buttons.items():
            if name == tab:
                btn.configure(
                    fg_color=T.SIDEBAR_BTN_ACTIVE_BG,
                    text_color=T.ACCENT_PRIMARY,
                    border_width=1,
                    border_color=T.ACCENT_PRIMARY,
                )
            else:
                btn.configure(
                    fg_color="transparent",
                    text_color=T.TEXT_PRIMARY,
                    border_width=0,
                )

        sw, variant = _parse_sw_variant(tab)
        self._sw_name = sw
        self._variant = variant

        self._sw_label.configure(text=sw or "—")
        self._variant_label.configure(text=variant or "—")

        self._status_label.configure(
            text=f"  Selected: {tab}   |   SW: {sw}   |   Variant: {variant}",
            text_color=T.ACCENT_SUCCESS,
        )

    # ── Folder browser ──────────────────────────────────────────
    def _browse_directory(self):
        directory = filedialog.askdirectory(
            title="Select Directory Containing PDF Reports",
        )
        if not directory:
            return
        self._pdf_dir = directory
        self._dir_entry.configure(state="normal")
        self._dir_entry.delete(0, "end")
        self._dir_entry.insert(0, directory)
        self._dir_entry.configure(state="disabled")
        if self._selected_tab:
            self._status_label.configure(
                text="  Ready. Click '▶  Start' to begin verification.",
                text_color=T.ACCENT_SUCCESS,
            )

    # ── Thread-safe GUI helpers ─────────────────────────────────
    def _ui(self, fn):
        self.after(0, fn)

    def _set_status(self, text: str, color: str = T.TEXT_PRIMARY):
        self._ui(lambda: self._status_label.configure(text=text, text_color=color))

    def _set_seg(self, idx: int, value: float):
        self._ui(lambda: self._progress.set_segment(idx, value))

    # ── Start analysis ──────────────────────────────────────────
    def _start_analysis(self):
        if not self._excel_file:
            self._status_label.configure(
                text="  Please select a SystemTestListe Excel file first.",
                text_color=T.ACCENT_WARNING,
            )
            return
        if not self._selected_tab:
            self._status_label.configure(
                text="  Please select a sheet / tab from the list.",
                text_color=T.ACCENT_WARNING,
            )
            return
        if not self._pdf_dir:
            self._status_label.configure(
                text="  Please select the PDF reports directory first.",
                text_color=T.ACCENT_WARNING,
            )
            return

        self._start_btn.configure(state="disabled", text="⏳  Analyzing...")
        self._progress.reset()
        threading.Thread(target=self._run_worker, daemon=True).start()

    # ── Background worker ───────────────────────────────────────
    def _run_worker(self):
        """Analysis worker – runs on a daemon thread."""
        from datetime import datetime
        import openpyxl as _openpyxl

        HEADER_CANDIDATES = [
            "ID", "Description", "Result", "Comment", "Problem Job",
            "Expert judgement", "Function", "Test Instance", "Safety",
            "ML", "Inspector", "Part of Concept",
        ]
        OUTPUT_COLS = ["ID", "Description", "Result"]

        # ── Step 1: Read selected sheet ─────────────────────────
        self._set_status(f"  Reading sheet '{self._selected_tab}' from Excel…")
        self._set_seg(0, 0.2)

        try:
            ext = os.path.splitext(self._excel_file)[1].lower()
            if ext == ".xls":
                import xlrd
                wb_in = xlrd.open_workbook(self._excel_file)
                ws_xls = wb_in.sheet_by_name(self._selected_tab)
                all_rows = [
                    [_cell_to_str(ws_xls.cell_value(r, c))
                     for c in range(ws_xls.ncols)]
                    for r in range(ws_xls.nrows)
                ]
            else:
                wb_in = _openpyxl.load_workbook(
                    self._excel_file, read_only=True, data_only=True
                )
                ws_xlsx = wb_in[self._selected_tab]
                all_rows = [
                    [_cell_to_str(v) for v in row]
                    for row in ws_xlsx.iter_rows(values_only=True)
                ]
                wb_in.close()
        except Exception as e:
            self._set_status(f"  Failed to read Excel: {e}", T.ACCENT_DANGER)
            self._ui(lambda: self._start_btn.configure(state="normal", text="▶  Start"))
            return

        self._set_seg(0, 0.6)

        # ── Find header row (most matching candidate columns) ───
        candidates_lower = [c.lower() for c in HEADER_CANDIDATES]
        best_idx, best_score = 0, 0
        for i, row in enumerate(all_rows):
            cells_lower = [str(c).strip().lower() for c in row]
            score = sum(1 for cand in candidates_lower if cand in cells_lower)
            if score > best_score:
                best_score, best_idx = score, i

        if best_score == 0:
            self._set_status(
                "  Could not find header row in the selected sheet.", T.ACCENT_DANGER
            )
            self._ui(lambda: self._start_btn.configure(state="normal", text="▶  Start"))
            return

        header_row = [str(c).strip() for c in all_rows[best_idx]]
        header_lower = [c.lower() for c in header_row]

        # Map output column names → source column indices
        col_map: dict[str, int | None] = {}
        for col_name in OUTPUT_COLS:
            try:
                col_map[col_name] = header_lower.index(col_name.lower())
            except ValueError:
                col_map[col_name] = None

        self._set_seg(0, 1.0)

        # ── Step 2: Filter rows where Description contains HILTS
        self._set_status("  Filtering HILTS rows…")
        self._set_seg(1, 0.3)

        desc_idx = col_map.get("Description")
        data_rows: list[list[str]] = []
        for row in all_rows[best_idx + 1:]:
            if not any(str(c).strip() for c in row):
                continue  # skip fully empty rows
            desc_val = (
                str(row[desc_idx]).strip()
                if desc_idx is not None and desc_idx < len(row)
                else ""
            )
            if "HILTS" in desc_val.upper():
                extracted = []
                for col_name in OUTPUT_COLS:
                    idx = col_map.get(col_name)
                    val = str(row[idx]).strip() if idx is not None and idx < len(row) else ""

                    # Result column: ensure only letters/specials (no digits)
                    if col_name == "Result" and val:
                        import re as _re
                        val = _re.sub(r"\d", "", val).strip()

                    extracted.append(val)
                data_rows.append(extracted)

        self._set_seg(1, 1.0)
        # ── Step 3: Create baseline STL_Helper Excel (part of Reading Excel)
        self._set_status(f"  Creating baseline STL_Helper Excel ({len(data_rows)} row(s))…")
        self._set_seg(0, 0.8)

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        output_dir = os.path.dirname(self._excel_file)
        output_path = os.path.join(output_dir, f"STL_Helper_{timestamp}.xlsx")

        try:
            wb_out = _openpyxl.Workbook()
            ws_out = wb_out.active
            ws_out.title = "MainSheet"
            out_cols = OUTPUT_COLS + ["PDFResult"]
            ws_out.append(out_cols)

            # write baseline rows (PDFResult empty)
            for r in data_rows:
                ws_out.append([r[0], r[1], r[2], ""])

            wb_out.save(output_path)
            self._set_seg(0, 1.0)
        except Exception as e:
            self._set_status(f"  Failed to create baseline output: {e}", T.ACCENT_DANGER)
            self._ui(lambda: self._start_btn.configure(state="normal", text="▶  Start"))
            return

        # ── Step 4: Scan PDFs and update the baseline file (Scanning PDFs segment)
        self._set_status(f"  Scanning PDFs and updating results…")
        self._set_seg(1, 0.0)

        # Build PDF file index (name without ext -> path)
        pdf_index: list[tuple[str, str]] = []
        for root, _dirs, files in os.walk(self._pdf_dir):
            for f in files:
                if f.lower().endswith('.pdf'):
                    name = os.path.splitext(f)[0]
                    pdf_index.append((name, os.path.join(root, f)))

        # helper: fuzzy match using difflib and process files in threads
        from difflib import SequenceMatcher
        import pdfplumber
        import concurrent.futures
        from concurrent.futures import as_completed

        KEYWORDS = ["passed", "failed", "error", "undefined", "not executed", "no result"]

        total = len(data_rows)
        processed = 0

        # worker: given (idx, row), return (idx, pdf_result, matched_path, score)
        def _process_row(args):
            idx, row = args
            descr = row[1] if len(row) > 1 else ""
            descr_norm = descr.strip().lower()
            best = (None, 0.0)
            for pname, ppath in pdf_index:
                score = SequenceMatcher(None, descr_norm, pname.lower()).ratio()
                if score > best[1]:
                    best = (ppath, score)

            pdf_result = "No report"
            matched = best[0]
            try:
                if matched and best[1] >= 0.95:
                    # choose page based on SW/Variant selection:
                    # - SW + Variant selected -> check page 3 (index 2)
                    # - SW selected -> check page 2 (index 1)
                    pref_idx = 1
                    try:
                        if getattr(self, "_sw_name", None) and getattr(self, "_variant", None):
                            pref_idx = 2
                        elif getattr(self, "_sw_name", None):
                            pref_idx = 1
                    except Exception:
                        pref_idx = 1

                    with pdfplumber.open(matched) as pdf:
                        txt = ""
                        if len(pdf.pages) > pref_idx:
                            txt = pdf.pages[pref_idx].extract_text() or ""
                        else:
                            # fallback to page 2 or page 1 if preferred not available
                            if len(pdf.pages) >= 2:
                                txt = pdf.pages[1].extract_text() or ""
                            elif len(pdf.pages) >= 1:
                                txt = pdf.pages[0].extract_text() or ""

                    t = (txt or "").lower()
                    found = None
                    for kw in KEYWORDS:
                        if kw in t:
                            found = kw
                            break

                    if getattr(self, "_sw_name", None) and getattr(self, "_variant", None):
                        # SW + Variant: return a data snippet from page 3 when no keyword found
                        if found:
                            pdf_result = found.title()
                        else:
                            # take first non-empty line as a short data summary
                            first_line = ""
                            for ln in (txt or "").splitlines():
                                s = ln.strip()
                                if s:
                                    first_line = s
                                    break
                            pdf_result = f"Page3: {first_line[:120]}" if first_line else "Page3: No data"
                    else:
                        # SW only (or default): check for pass/fail keywords
                        if found:
                            pdf_result = found.title()
                        else:
                            pdf_result = "No Result"
                else:
                    pdf_result = "No report"
            except Exception:
                pdf_result = "No Result"

            return (idx, pdf_result, matched, best[1])

        # Process names one-by-one and update progress per name
        processed = 0
        for idx, row in enumerate(data_rows):
            descr = row[1] if len(row) > 1 else ""
            descr_norm = descr.strip().lower()
            best = (None, 0.0)
            for pname, ppath in pdf_index:
                score = SequenceMatcher(None, descr_norm, pname.lower()).ratio()
                if score > best[1]:
                    best = (ppath, score)

            pdf_result = "No report"
            matched = best[0]
            try:
                if matched and best[1] >= 0.95:
                    pref_idx = 1
                    try:
                        if getattr(self, "_sw_name", None) and getattr(self, "_variant", None):
                            pref_idx = 2
                        elif getattr(self, "_sw_name", None):
                            pref_idx = 1
                    except Exception:
                        pref_idx = 1

                    with pdfplumber.open(matched) as pdf:
                        txt = ""
                        if len(pdf.pages) > pref_idx:
                            txt = pdf.pages[pref_idx].extract_text() or ""
                        else:
                            if len(pdf.pages) >= 2:
                                txt = pdf.pages[1].extract_text() or ""
                            elif len(pdf.pages) >= 1:
                                txt = pdf.pages[0].extract_text() or ""

                    t = (txt or "").lower()
                    found = None
                    for kw in KEYWORDS:
                        if kw in t:
                            found = kw
                            break

                    if getattr(self, "_sw_name", None) and getattr(self, "_variant", None):
                        if found:
                            pdf_result = found.title()
                        else:
                            first_line = ""
                            for ln in (txt or "").splitlines():
                                s = ln.strip()
                                if s:
                                    first_line = s
                                    break
                            pdf_result = f"Page3: {first_line[:120]}" if first_line else "Page3: No data"
                    else:
                        if found:
                            pdf_result = found.title()
                        else:
                            pdf_result = "No Result"
                else:
                    pdf_result = "No report"
            except Exception:
                pdf_result = "No Result"

            # update the existing workbook row: header at row 1, data starts row 2
            out_row = 2 + idx
            try:
                ws_out.cell(row=out_row, column=4, value=pdf_result)
            except Exception:
                pass

            processed = idx + 1
            frac = processed / total if total else 1.0
            self._set_seg(1, frac)
            self._set_status(
                f"  Scanning PDFs {processed}/{total}  |  SW: {self._sw_name or '—'}  |  Variant: {self._variant or '—'}  |  Last: {pdf_result}",
                T.TEXT_SECONDARY,
            )

        # save final workbook with PDF results
        try:
            wb_out.save(output_path)
        except Exception as e:
            self._set_status(f"  Failed to save updated output: {e}", T.ACCENT_DANGER)
            self._ui(lambda: self._start_btn.configure(state="normal", text="▶  Start"))
            return

        self._result_path = output_path
        self._set_seg(2, 1.0)

        final = (
            f"  ✅  Done – {len(data_rows)} HILTS row(s) written to "
            f"{os.path.basename(output_path)}  |  Header found at row {best_idx + 1}"
        )
        self._set_status(final, T.ACCENT_SUCCESS)
        self._ui(lambda: self._start_btn.configure(state="normal", text="▶  Start"))

